"""
=============================================================
  METROMECANICA — Caracterizacion de Balanza
  RADWAG AS 82/220.X2 | BIOBASE BL-5000 | WANT GT-30000TR
  NMP 004:2007 — Clase M2 | Procedimiento ABA
  s(DI) 15 decimales | Varianza | Tendencia
  ISO/IEC 17025 | v4.0
=============================================================
  pip install pyserial matplotlib openpyxl pyttsx3
=============================================================
  v4.0:
  - Subpestanas por pesa en cada balanza
  - WANT: 10,15,20,25 kg | BIOBASE: 1,2,5 kg | RADWAG: 1-200g
  - Ingreso manual WANT con display grande (22pt)
  - Coma decimal INACAL en toda la interfaz
  - Excel: una hoja por ciclo, formato referencia
=============================================================
"""

import serial
import serial.tools.list_ports
import socket
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json, os, re, math, threading, time, sys, subprocess
from datetime import datetime

try:
    import pyttsx3
    VOZ_OK = True
except ImportError:
    VOZ_OK = False

import matplotlib
matplotlib.use("TkAgg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

DIR_APP   = os.path.dirname(os.path.abspath(__file__))
FILE_HIST = os.path.join(DIR_APP, "historial_carac.json")

PESAS_RADWAG = [
    (1.0,   "1 g",    1.5,   4),
    (2.0,   "2 g",    3.0,   4),
    (5.0,   "5 g",    5.0,   4),
    (10.0,  "10 g",   10.0,  4),
    (20.0,  "20 g",   25.0,  4),
    (50.0,  "50 g",   50.0,  4),
    (100.0, "100 g",  100.0, 4),
    (200.0, "200 g",  150.0, 4),
]
PESAS_BIOBASE = [
    (1000.0, "1 kg",  500.0,  2),
    (2000.0, "2 kg",  1000.0, 2),
    (5000.0, "5 kg",  2500.0, 2),
]
PESAS_WANT = [
    (10000.0, "10 kg", 5000.0,  1),
    (15000.0, "15 kg", 7500.0,  1),
    (20000.0, "20 kg", 10000.0, 1),
    (25000.0, "25 kg", 12500.0, 1),
]

N_LECTURAS  = 10
N_CICLOS    = 10
RADWAG_IP   = "192.168.18.65"
RADWAG_PORT = 4001

BG      = "#080d18"; PANEL   = "#0f1828"; PANEL2  = "#141f2e"
BORDER  = "#1a2940"; ACCENT  = "#00c8e0"; ACCENT2 = "#0077b6"
GREEN   = "#22c55e"; RED     = "#ef4444"; YELLOW  = "#f59e0b"
TXT     = "#cdd9e5"; TXT_DIM = "#4a6480"; TEAL    = "#0d9488"
PURPLE  = "#7c3aed"; ORANGE  = "#f97316"
FN_UI   = ("Georgia", 9)
FN_SM   = ("Georgia", 8)
FN_BOLD = ("Georgia", 9, "bold")
ORDINAL = {1:"primer",2:"segundo",3:"tercer",4:"cuarto",
           5:"quinto",6:"sexto",7:"septimo",8:"octavo",
           9:"noveno",10:"decimo"}


def fmt_coma(v, d):
    return format(v, f".{d}f").replace(".", ",")

def fmt_stat(v, d=15):
    return format(v, f".{d}f").replace(".", ",")

def fmt_emp(emp_mg):
    return f"{emp_mg/1000:g} g" if emp_mg >= 1000 else f"{emp_mg:g} mg"

def parsear_serial(raw):
    m = re.search(r'([+-]?\s*\d+\.?\d*)\s*(g|kg)', raw, re.I)
    if m:
        try:
            v = float(m.group(1).replace(" ", ""))
            return v*1000 if m.group(2).lower()=="kg" else v
        except: pass
    return None

def parsear_radwag(raw):
    m = re.search(r'([+-]?\s*\d+\.?\d*)\s*g', raw, re.I)
    if m:
        try: return float(m.group(1).replace(" ", ""))
        except: pass
    return None

def delta_i(ir1, it, ir2):
    return it - (ir1+ir2)/2.0

def s_delta(lecturas):
    if len(lecturas) < 2: return 0.0
    deltas = [delta_i(l["ir1"],l["it"],l["ir2"]) for l in lecturas]
    n=len(deltas); mean=sum(deltas)/n
    return math.sqrt(sum((d-mean)**2 for d in deltas)/(n-1))

def varianza_ciclos(sdis):
    n=len(sdis)
    if n<2: return None
    mean=sum(sdis)/n
    return sum((s-mean)**2 for s in sdis)/(n-1)

def cargar_hist():
    try:
        with open(FILE_HIST,"r",encoding="utf-8") as f:
            return json.load(f)
    except: return {}

def guardar_hist(h):
    with open(FILE_HIST,"w",encoding="utf-8") as f:
        json.dump(h,f,indent=2,ensure_ascii=False)

def hablar(texto):
    if not VOZ_OK: return
    def _run():
        try:
            eng=pyttsx3.init(); eng.setProperty("rate",150)
            for v in eng.getProperty("voices"):
                if "spanish" in v.name.lower() or "es" in v.id.lower():
                    eng.setProperty("voice",v.id); break
            eng.say(texto); eng.runAndWait(); eng.stop()
        except: pass
    threading.Thread(target=_run,daemon=True).start()

def abrir_archivo(ruta):
    if sys.platform=="win32": os.startfile(ruta)
    elif sys.platform=="darwin": subprocess.call(["open",ruta])
    else: subprocess.call(["xdg-open",ruta])


class ConexionSerial:
    def __init__(self,on_dato):
        self.on_dato=on_dato; self.ser=None; self.activo=False
    def conectar(self,puerto,baud=9600):
        try:
            self.ser=serial.Serial(port=puerto,baudrate=baud,
                bytesize=serial.EIGHTBITS,parity=serial.PARITY_NONE,
                stopbits=serial.STOPBITS_ONE,timeout=3)
            self.activo=True
            threading.Thread(target=self._loop,daemon=True).start()
            return True
        except: return False
    def desconectar(self):
        self.activo=False
        if self.ser:
            try: self.ser.close()
            except: pass
    def _loop(self):
        while self.activo:
            try:
                if self.ser and self.ser.in_waiting>0:
                    raw=self.ser.readline().decode("ascii",errors="ignore").strip()
                    if raw:
                        val=parsear_serial(raw)
                        if val is not None: self.on_dato(val,raw)
                time.sleep(0.05)
            except: break


class ConexionRadwag:
    def __init__(self,on_dato):
        self.on_dato=on_dato; self.sock=None
        self.activo=False; self.ip=RADWAG_IP; self.port=RADWAG_PORT
    def conectar(self,ip=None,port=None):
        if ip: self.ip=ip
        if port: self.port=port
        self.activo=True
        threading.Thread(target=self._loop,daemon=True).start()
        return True
    def desconectar(self):
        self.activo=False
        if self.sock:
            try: self.sock.close()
            except: pass
            self.sock=None
    def _conectar_sock(self):
        if self.sock:
            try: self.sock.close()
            except: pass
            self.sock=None
        try:
            s=socket.socket(socket.AF_INET,socket.SOCK_STREAM)
            s.setsockopt(socket.SOL_SOCKET,socket.SO_REUSEADDR,1)
            s.settimeout(5); s.connect((self.ip,self.port))
            s.settimeout(2); self.sock=s; return True
        except: return False
    def _loop(self):
        while self.activo:
            if not self._conectar_sock():
                time.sleep(3); continue
            self.on_dato(None,"__CONNECTED__")
            buf=""
            while self.activo:
                try:
                    data=self.sock.recv(256)
                    if not data: break
                    buf+=data.decode("ascii",errors="ignore")
                    while "\r\n" in buf:
                        linea,buf=buf.split("\r\n",1)
                        linea=linea.strip()
                        if linea:
                            val=parsear_radwag(linea)
                            if val is not None: self.on_dato(val,linea)
                except socket.timeout: continue
                except: break
            if self.activo:
                self.on_dato(None,"__DISCONNECTED__"); time.sleep(2)


class PanelABA(tk.Frame):
    def __init__(self,parent,pesa_cfg,balanza_cfg,
                 hist_ref,on_completado,cx_getter,**kw):
        super().__init__(parent,bg=BG,**kw)
        self.pesa=pesa_cfg; self.bcfg=balanza_cfg
        self.hist=hist_ref; self.on_completado=on_completado
        self.cx_getter=cx_getter
        self.paso=0; self.lecturas=[]; self.tmp_ir1=None
        self.tmp_it=None; self.ultimo_sdi=None
        self._build()

    @property
    def key(self): return f"{self.bcfg['nombre']}_{self.pesa[1]}"
    @property
    def dd(self): return self.pesa[3]

    def _card(self,parent,titulo,color=None,expand=False):
        c=color or ACCENT
        outer=tk.Frame(parent,bg=BORDER)
        outer.pack(fill="both" if expand else "x",expand=expand,pady=(0,4))
        tk.Frame(outer,bg=c,width=3).pack(side="left",fill="y")
        inner=tk.Frame(outer,bg=PANEL,padx=8,pady=6)
        inner.pack(fill="both",expand=True)
        tk.Label(inner,text=titulo.upper(),bg=PANEL,fg=c,
                 font=("Georgia",7,"bold")).pack(anchor="w")
        tk.Frame(inner,bg=BORDER,height=1).pack(fill="x",pady=(2,5))
        return inner

    def _build(self):
        color=self.bcfg["color"]
        body=tk.Frame(self,bg=BG)
        body.pack(fill="both",expand=True,padx=6,pady=4)
        col_izq=tk.Frame(body,bg=BG,width=400)
        col_izq.pack(side="left",fill="y",padx=(0,5))
        col_izq.pack_propagate(False)
        col_der=tk.Frame(body,bg=BG)
        col_der.pack(side="right",fill="both",expand=True)
        self._build_ciclo(col_izq,color)
        self._build_historial(col_der,color)
        self._build_grafico(col_der)

    def _build_ciclo(self,parent,color):
        p=self._card(parent,
            f"Ciclo ABA  {self.pesa[1]}  (EMP +/-{fmt_emp(self.pesa[2])})",
            color)
        self.lbl_paso=tk.Label(p,
            text="Presiona Iniciar nueva caracterizacion",
            bg=PANEL,fg=TXT_DIM,
            font=("Courier New",8,"bold"),
            wraplength=370,justify="left")
        self.lbl_paso.pack(anchor="w",pady=(0,4))

        tk.Frame(p,bg=BORDER,height=1).pack(fill="x",pady=(0,4))
        tk.Label(p,text="DATOS PESA PATRON",bg=PANEL,fg=TEAL,
                 font=("Georgia",7,"bold")).pack(anchor="w")

        def campo(lbl,attr,val=""):
            row=tk.Frame(p,bg=PANEL); row.pack(fill="x",pady=1)
            tk.Label(row,text=lbl,bg=PANEL,fg=TXT,
                     font=FN_UI,width=14,anchor="w").pack(side="left")
            e=tk.Entry(row,font=("Courier New",9),bg=PANEL2,
                       fg=TXT,insertbackground=TEAL,
                       relief="flat",bd=2,width=18)
            e.insert(0,val); e.pack(side="left",padx=4)
            setattr(self,attr,e)

        campo("ID / Codigo:","e_pat_id")
        campo("N Certificado:","e_pat_cert")
        campo("Fecha Cal.:","e_pat_fecha",
              datetime.now().strftime("%Y-%m-%d"))

        tk.Frame(p,bg=BORDER,height=1).pack(fill="x",pady=(5,4))
        self.lbl_ciclos_info=tk.Label(p,text="Ciclos guardados: 0",
            bg=PANEL,fg=TXT_DIM,font=FN_SM)
        self.lbl_ciclos_info.pack(anchor="w")
        self.lbl_var_hist=tk.Label(p,
            text="Varianza: --- (min. 2 ciclos)",
            bg=PANEL,fg=TXT_DIM,
            font=("Courier New",8,"bold"))
        self.lbl_var_hist.pack(anchor="w",pady=(0,4))

        tk.Button(p,text="  Iniciar nueva caracterizacion  ",
            bg=TEAL,fg="white",font=FN_BOLD,
            relief="flat",padx=10,pady=5,
            command=self._iniciar).pack(fill="x",pady=(0,6))

        parcial=tk.Frame(p,bg="#0a1525",padx=8,pady=6)
        parcial.pack(fill="x",pady=(0,4))
        fila_p=tk.Frame(parcial,bg="#0a1525"); fila_p.pack(fill="x")
        for txt,attr in [("Ir1:","lbl_ir1"),("It:","lbl_it"),("Ir2:","lbl_ir2")]:
            tk.Label(fila_p,text=txt,bg="#0a1525",fg=TXT_DIM,
                     font=FN_SM).pack(side="left",padx=(0,2))
            lv=tk.Label(fila_p,text="---",bg="#0a1525",fg=TXT_DIM,
                        font=("Courier New",9,"bold"),width=10)
            lv.pack(side="left",padx=(0,8))
            setattr(self,attr,lv)
        self.lbl_di=tk.Label(parcial,text="DI: ---",
            bg="#0a1525",fg=ACCENT,font=("Courier New",8,"bold"))
        self.lbl_di.pack(anchor="w",pady=(4,0))

        self.lbl_n=tk.Label(p,text=f"0 / {N_LECTURAS} lecturas",
            bg=PANEL,fg=TXT_DIM,font=("Courier New",8))
        self.lbl_n.pack(anchor="w",pady=(0,3))

        cols=("N","PATRON Ir1","CALIBRAR It","PATRON Ir2","DI")
        self.tbl=ttk.Treeview(p,columns=cols,show="headings",height=6)
        for col,w in zip(cols,[28,95,95,95,80]):
            self.tbl.heading(col,text=col)
            self.tbl.column(col,width=w,anchor="center",minwidth=w)
        sb=ttk.Scrollbar(p,orient="vertical",command=self.tbl.yview)
        sb.pack(side="right",fill="y")
        self.tbl.configure(yscrollcommand=sb.set)
        self.tbl.pack(fill="x")

        tk.Frame(p,bg=BORDER,height=1).pack(fill="x",pady=(6,3))

        row_s=tk.Frame(p,bg=PANEL); row_s.pack(fill="x",pady=2)
        tk.Label(row_s,text="s(DI):",bg=PANEL,fg=TXT,
                 font=("Georgia",9,"bold"),
                 width=12,anchor="e").pack(side="left")
        self.lbl_sdi=tk.Label(row_s,text="---",
            bg=PANEL2,fg=ACCENT,
            font=("Courier New",9,"bold"),
            anchor="w",padx=8,pady=3,relief="flat",width=28)
        self.lbl_sdi.pack(side="left",padx=4)
        tk.Label(row_s,text="g",bg=PANEL,fg=TXT_DIM,font=FN_SM).pack(side="left")

        row_v=tk.Frame(p,bg=PANEL); row_v.pack(fill="x",pady=2)
        tk.Label(row_v,text="Varianza:",bg=PANEL,fg=TXT,
                 font=("Georgia",9,"bold"),
                 width=12,anchor="e").pack(side="left")
        self.lbl_var=tk.Label(row_v,text="---",
            bg=TEAL,fg="white",
            font=("Courier New",9,"bold"),
            anchor="w",padx=8,pady=3,relief="flat",width=28)
        self.lbl_var.pack(side="left",padx=4)
        tk.Label(row_v,text="g^2  (>=2 ciclos)",
                 bg=PANEL,fg=TXT_DIM,font=FN_SM).pack(side="left")

    def _build_historial(self,parent,color):
        p=self._card(parent,"Historial de ciclos",color)
        cols=("N","Fecha","ID Patron","N Cert.","s(DI) g","Var g2")
        self.tbl_hist=ttk.Treeview(p,columns=cols,show="headings",height=4)
        for col,w in zip(cols,[28,140,90,90,150,150]):
            self.tbl_hist.heading(col,text=col)
            self.tbl_hist.column(col,width=w,anchor="center",minwidth=28)
        sy=ttk.Scrollbar(p,orient="vertical",command=self.tbl_hist.yview)
        sx=ttk.Scrollbar(p,orient="horizontal",command=self.tbl_hist.xview)
        self.tbl_hist.configure(yscrollcommand=sy.set,xscrollcommand=sx.set)
        sy.pack(side="right",fill="y")
        self.tbl_hist.pack(fill="both",expand=True)
        sx.pack(fill="x")

    def _build_grafico(self,parent):
        p=self._card(parent,f"Tendencia s(DI)  {self.pesa[1]}",
                     PURPLE,expand=True)
        self.fig=Figure(figsize=(5,2.4),facecolor="#0f1828")
        self.ax=self.fig.add_subplot(111)
        self._estilo_ax()
        self.canvas=FigureCanvasTkAgg(self.fig,master=p)
        self.canvas.get_tk_widget().pack(fill="both",expand=True)

    def _estilo_ax(self):
        self.ax.clear(); self.ax.set_facecolor("#0a1525")
        self.ax.tick_params(colors="#4a6480",labelsize=8)
        for sp in ["bottom","left"]:
            self.ax.spines[sp].set_color("#1a2940")
        self.ax.spines["top"].set_visible(False)
        self.ax.spines["right"].set_visible(False)
        self.ax.set_xlabel("Ciclo",color="#4a6480",fontsize=8)
        self.ax.set_ylabel("s(DI) g",color="#4a6480",fontsize=8)
        self.ax.grid(color="#1a2940",linestyle="--",linewidth=0.5)

    def _iniciar(self):
        if self.bcfg["tipo"]!="manual":
            cx=self.cx_getter()
            if not cx or not getattr(cx,"activo",False):
                messagebox.showwarning("Sin conexion",
                    "Conecta la balanza primero.")
                return
        self.paso=1; self.lecturas=[]
        self.tmp_ir1=None; self.tmp_it=None; self.ultimo_sdi=None
        for i in self.tbl.get_children(): self.tbl.delete(i)
        for a in ("lbl_ir1","lbl_it","lbl_ir2"):
            getattr(self,a).config(text="---",fg=TXT_DIM)
        self.lbl_di.config(text="DI: ---",fg=ACCENT)
        self.lbl_sdi.config(text="---",fg=ACCENT)
        self.lbl_var.config(text="---",fg="white")
        self.lbl_n.config(text=f"0 / {N_LECTURAS} lecturas",fg=TXT_DIM)
        self._upd_paso()

    def _upd_paso(self):
        n=len(self.lecturas)+1
        msgs={
            1:f">> Lect. {n}/{N_LECTURAS}  Ir1: PATRON -> PRINT",
            2:f">> Lect. {n}/{N_LECTURAS}  It:  CALIBRAR -> PRINT",
            3:f">> Lect. {n}/{N_LECTURAS}  Ir2: PATRON -> PRINT",
        }
        self.lbl_paso.config(
            text=msgs.get(self.paso,"Presiona Iniciar nueva caracterizacion"),
            fg=YELLOW if self.paso>0 else TXT_DIM)

    def recibir(self,valor):
        if self.paso==0: return
        if self.paso==1:
            self.tmp_ir1=valor; self.paso=2
            self.lbl_ir1.config(text=fmt_coma(valor,self.dd)+" g",fg=GREEN)
            self.lbl_it.config(text="---",fg=TXT_DIM)
            self.lbl_ir2.config(text="---",fg=TXT_DIM)
            self._upd_paso()
        elif self.paso==2:
            self.tmp_it=valor; self.paso=3
            self.lbl_it.config(text=fmt_coma(valor,self.dd)+" g",fg=GREEN)
            self.lbl_ir2.config(text="esperando...",fg=YELLOW)
            self._upd_paso()
        elif self.paso==3:
            ir1=self.tmp_ir1; it=self.tmp_it; ir2=valor
            di=delta_i(ir1,it,ir2)
            self.lbl_ir2.config(text=fmt_coma(ir2,self.dd)+" g",fg=GREEN)
            self.lbl_di.config(text=f"DI = {fmt_stat(di)} g",fg=ACCENT)
            self.lecturas.append({"ir1":ir1,"it":it,"ir2":ir2})
            n=len(self.lecturas)
            self.tbl.insert("","end",values=(
                n,
                fmt_coma(ir1,self.dd),
                fmt_coma(it,self.dd),
                fmt_coma(ir2,self.dd),
                fmt_coma(di,min(self.dd,3))))
            hijos=self.tbl.get_children()
            if hijos: self.tbl.see(hijos[-1])
            self.lbl_n.config(text=f"{n} / {N_LECTURAS} lecturas",fg=TXT_DIM)
            if n>=2:
                sdi_p=s_delta(self.lecturas)
                self.lbl_sdi.config(text=fmt_stat(sdi_p)+" g",fg=YELLOW)
                deltas=[delta_i(l["ir1"],l["it"],l["ir2"]) for l in self.lecturas]
                mn=sum(deltas)/len(deltas)
                var=sum((d-mn)**2 for d in deltas)/(len(deltas)-1)
                self.lbl_var.config(text=fmt_stat(var)+" g^2")
            self.tmp_ir1=None; self.tmp_it=None
            for a in ("lbl_ir1","lbl_it","lbl_ir2"):
                getattr(self,a).config(text="---",fg=TXT_DIM)
            if n>=N_LECTURAS: self._completar()
            else: self.paso=1; self._upd_paso()

    def _completar(self):
        self.paso=0
        sdi=s_delta(self.lecturas); self.ultimo_sdi=sdi
        deltas=[delta_i(l["ir1"],l["it"],l["ir2"]) for l in self.lecturas]
        mn=sum(deltas)/len(deltas)
        var=sum((d-mn)**2 for d in deltas)/(len(deltas)-1)
        self.lbl_sdi.config(text=fmt_stat(sdi)+" g",fg=GREEN)
        self.lbl_var.config(text=fmt_stat(var)+" g^2",fg="white",bg=TEAL)
        n_sig=len(self.hist.get(self.key,{}).get("ciclos",[]))+1
        self.lbl_n.config(text=f"OK  {N_LECTURAS}/{N_LECTURAS} completo",fg=GREEN)
        self.lbl_paso.config(
            text=f"OK  {ORDINAL.get(n_sig,str(n_sig))} ciclo completado\n"
                 f"   s(DI) = {fmt_stat(sdi)} g\n"
                 f"   Presiona GUARDAR CICLO",
            fg=GREEN)
        if self.on_completado: self.on_completado(self)
        hablar(f"{ORDINAL.get(n_sig,str(n_sig))} ciclo completo")

    def guardar_ciclo(self):
        if self.ultimo_sdi is None: return
        ts=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if self.key not in self.hist: self.hist[self.key]={"ciclos":[]}
        ciclos=self.hist[self.key]["ciclos"]
        n_ciclo=len(ciclos)+1
        ciclos.append({"n":n_ciclo,"fecha":ts,"sdi":self.ultimo_sdi,
            "pat_id":self.e_pat_id.get().strip(),
            "pat_cert":self.e_pat_cert.get().strip(),
            "pat_fecha":self.e_pat_fecha.get().strip(),
            "lecturas":list(self.lecturas)})
        guardar_hist(self.hist)
        self._upd_info(); self._upd_hist(); self._upd_graf()
        total=len(ciclos)
        if total>=N_CICLOS:
            hablar(f"Caracterizacion terminada. Pesa {self.pesa[1]}.")
            messagebox.showinfo("Terminada",
                f"Completada!\nPesa: {self.pesa[1]}\n"
                f"Ciclos: {total}\n"
                f"s(DI) = {fmt_stat(self.ultimo_sdi)} g")
        else:
            hablar(f"Ciclo {n_ciclo} guardado")
            messagebox.showinfo("Guardado",
                f"Ciclo {n_ciclo} guardado.\n"
                f"s(DI) = {fmt_stat(self.ultimo_sdi)} g")
        self.ultimo_sdi=None

    def _upd_info(self):
        ciclos=self.hist.get(self.key,{}).get("ciclos",[])
        n=len(ciclos)
        self.lbl_ciclos_info.config(text=f"Ciclos guardados: {n}")
        if n>=2:
            sdis=[c["sdi"] for c in ciclos]
            var=varianza_ciclos(sdis)
            self.lbl_var_hist.config(
                text=f"Varianza: {fmt_stat(var)} g^2",
                fg=GREEN if var<1e-4 else YELLOW)
        else:
            self.lbl_var_hist.config(
                text="Varianza: --- (min. 2 ciclos)",fg=TXT_DIM)

    def _upd_hist(self):
        for i in self.tbl_hist.get_children(): self.tbl_hist.delete(i)
        ciclos=self.hist.get(self.key,{}).get("ciclos",[])
        sdis=[c["sdi"] for c in ciclos]
        for i,c in enumerate(ciclos):
            var_s="---"
            if i>=1:
                vac=varianza_ciclos(sdis[:i+1])
                if vac is not None: var_s=fmt_stat(vac)
            self.tbl_hist.insert("","end",values=(
                c["n"],c["fecha"],
                c.get("pat_id","---"),c.get("pat_cert","---"),
                fmt_stat(c["sdi"]),var_s))

    def _upd_graf(self):
        ciclos=self.hist.get(self.key,{}).get("ciclos",[])
        self._estilo_ax()
        if not ciclos:
            self.ax.text(0.5,0.5,"Sin ciclos",
                transform=self.ax.transAxes,
                ha="center",va="center",color="#4a6480",fontsize=9)
            self.canvas.draw(); return
        xs=[c["n"] for c in ciclos]; ys=[c["sdi"] for c in ciclos]
        self.ax.plot(xs,ys,color="#00c8e0",linewidth=1.5,
            marker="o",markersize=5,
            markerfacecolor="#22c55e",markeredgecolor="#22c55e")
        if len(ys)>=2:
            media=sum(ys)/len(ys)
            var=varianza_ciclos(ys)
            sigma=math.sqrt(var) if var else 0
            self.ax.axhline(media,color="#f59e0b",
                linewidth=1,linestyle="--",label="media")
            self.ax.axhline(media+sigma,color="#ef4444",
                linewidth=0.8,linestyle=":",alpha=0.7,label="+s")
            self.ax.axhline(media-sigma,color="#ef4444",
                linewidth=0.8,linestyle=":",alpha=0.7)
        if xs:
            self.ax.annotate(f"{ys[-1]:.6f}",(xs[-1],ys[-1]),
                textcoords="offset points",xytext=(6,6),
                fontsize=7,color="#00c8e0")
        self.ax.set_title(
            f"s(DI)  {self.bcfg['nombre']}  {self.pesa[1]}",
            color="#cdd9e5",fontsize=8,pad=4)
        if len(ys)>=2:
            self.ax.legend(fontsize=7,labelcolor="#cdd9e5",
                facecolor="#0a1525",edgecolor="#1a2940",loc="upper right")
        self.ax.set_xticks(xs)
        self.fig.tight_layout(pad=1.0)
        self.canvas.draw()

    def refresh(self):
        self._upd_info(); self._upd_hist(); self._upd_graf()


class PanelBalanza(tk.Frame):
    def __init__(self,parent,cfg,hist_ref,on_completado,**kw):
        super().__init__(parent,bg=BG,**kw)
        self.cfg=cfg; self.hist=hist_ref
        self.on_completado=on_completado
        self.cx=None; self.panels_aba={}; self.tab_activa=None
        self._build()

    def get_cx(self): return self.cx

    def _build(self):
        color=self.cfg["color"]; tipo=self.cfg["tipo"]
        franja=tk.Frame(self,bg=PANEL,padx=10,pady=7)
        franja.pack(fill="x")
        tk.Label(franja,text=f"  {self.cfg['nombre']}",
                 bg=PANEL,fg=color,
                 font=("Georgia",10,"bold")).pack(side="left",padx=(0,16))

        if tipo=="serial":
            tk.Label(franja,text="Puerto:",bg=PANEL,fg=TXT,font=FN_UI).pack(side="left")
            self.combo_port=ttk.Combobox(franja,width=7,state="readonly")
            puertos=[x.device for x in serial.tools.list_ports.comports()]
            self.combo_port["values"]=puertos
            dflt=self.cfg.get("puerto","COM6")
            self.combo_port.set(dflt if dflt in puertos else (puertos[0] if puertos else ""))
            self.combo_port.pack(side="left",padx=4)
            tk.Label(franja,text="Baud:",bg=PANEL,fg=TXT,font=FN_UI).pack(side="left")
            self.combo_baud=ttk.Combobox(franja,width=6,state="readonly",
                values=["9600","19200","4800","2400"])
            self.combo_baud.set(str(self.cfg.get("baud",9600)))
            self.combo_baud.pack(side="left",padx=4)
            tk.Button(franja,text="R",bg=PANEL2,fg=TXT_DIM,
                      font=("Georgia",10),relief="flat",
                      command=self._refresh_ports).pack(side="left",padx=2)
            self.btn_cx=tk.Button(franja,text="Conectar",bg=color,fg="white",
                font=FN_BOLD,relief="flat",padx=10,pady=2,command=self._toggle_cx)
            self.btn_cx.pack(side="left",padx=8)
            self.lbl_cx=tk.Label(franja,text="Desconectado",bg=PANEL,fg=RED,font=FN_SM)
            self.lbl_cx.pack(side="left")

        elif tipo=="wifi":
            tk.Label(franja,text="IP:",bg=PANEL,fg=TXT,font=FN_UI).pack(side="left")
            self.e_ip=tk.Entry(franja,width=14,font=("Courier New",9),
                bg=PANEL2,fg=TXT,insertbackground=color,relief="flat",bd=2)
            self.e_ip.insert(0,self.cfg.get("ip",RADWAG_IP))
            self.e_ip.pack(side="left",padx=4)
            tk.Label(franja,text="Puerto:",bg=PANEL,fg=TXT,font=FN_UI).pack(side="left")
            self.e_wport=tk.Entry(franja,width=5,font=("Courier New",9),
                bg=PANEL2,fg=TXT,insertbackground=color,relief="flat",bd=2)
            self.e_wport.insert(0,str(self.cfg.get("port",RADWAG_PORT)))
            self.e_wport.pack(side="left",padx=4)
            self.btn_cx=tk.Button(franja,text="Conectar",bg=color,fg="white",
                font=FN_BOLD,relief="flat",padx=10,pady=2,command=self._toggle_cx)
            self.btn_cx.pack(side="left",padx=8)
            self.lbl_cx=tk.Label(franja,text="Desconectado",bg=PANEL,fg=RED,font=FN_SM)
            self.lbl_cx.pack(side="left")

        elif tipo=="manual":
            tk.Label(franja,text="Valor leido:",bg=PANEL,fg=TXT,font=FN_UI).pack(side="left")
            self.e_manual=tk.Entry(franja,width=11,
                font=("Courier New",22,"bold"),
                bg="#0a1525",fg=GREEN,
                insertbackground=color,
                relief="flat",bd=3,justify="right")
            self.e_manual.pack(side="left",padx=6,ipady=2)
            tk.Label(franja,text="g",bg=PANEL,fg=TXT_DIM,
                     font=("Courier New",16)).pack(side="left")
            tk.Button(franja,text="  REGISTRAR  ",
                bg=color,fg="white",
                font=("Georgia",11,"bold"),
                relief="flat",padx=10,pady=4,
                command=self._registrar_manual).pack(side="left",padx=10)
            self.e_manual.bind("<Return>",lambda e:self._registrar_manual())
            self.lbl_cx=tk.Label(franja,text="Modo ingreso manual",
                bg=PANEL,fg=ORANGE,font=FN_SM)
            self.lbl_cx.pack(side="left")

        tk.Frame(self,bg=color,height=2).pack(fill="x")

        nav=tk.Frame(self,bg=PANEL2,height=30)
        nav.pack(fill="x"); nav.pack_propagate(False)
        content=tk.Frame(self,bg=BG)
        content.pack(fill="both",expand=True)
        self.tab_frames={}; self.tab_btns={}

        for pesa in self.cfg["pesas"]:
            lbl=pesa[1]
            frame=tk.Frame(content,bg=BG)
            self.tab_frames[lbl]=frame
            pa=PanelABA(frame,pesa,self.cfg,self.hist,
                        self.on_completado,cx_getter=self.get_cx)
            pa.pack(fill="both",expand=True)
            self.panels_aba[lbl]=pa
            btn=tk.Button(nav,text=f"  {lbl}  ",
                bg=PANEL2,fg=TXT_DIM,font=("Georgia",8),
                relief="flat",padx=6,pady=4,
                command=lambda k=lbl:self._switch_pesa(k))
            btn.pack(side="left"); self.tab_btns[lbl]=btn

        self._switch_pesa(self.cfg["pesas"][0][1])

    def _switch_pesa(self,lbl):
        color=self.cfg["color"]
        for k,frame in self.tab_frames.items():
            frame.pack_forget()
            self.tab_btns[k].config(bg=PANEL2,fg=TXT_DIM,font=("Georgia",8))
        self.tab_frames[lbl].pack(fill="both",expand=True)
        self.tab_btns[lbl].config(bg=color,fg="white",font=("Georgia",8,"bold"))
        self.tab_activa=lbl; self.panels_aba[lbl].refresh()

    def panel_activo(self):
        return self.panels_aba.get(self.tab_activa)

    def _refresh_ports(self):
        p=[x.device for x in serial.tools.list_ports.comports()]
        self.combo_port["values"]=p

    def _toggle_cx(self):
        color=self.cfg["color"]; tipo=self.cfg["tipo"]
        if self.cx and getattr(self.cx,"activo",False):
            self.cx.desconectar(); self.cx=None
            self.btn_cx.config(text="Conectar",bg=color)
            self.lbl_cx.config(text="Desconectado",fg=RED)
        else:
            if tipo=="serial":
                puerto=self.combo_port.get()
                baud=int(self.combo_baud.get())
                self.cx=ConexionSerial(on_dato=self._on_dato)
                ok=self.cx.conectar(puerto,baud)
                if ok:
                    self.btn_cx.config(text="Desconectar",bg=RED)
                    self.lbl_cx.config(
                        text=f"Conectado {puerto} @ {baud}",fg=GREEN)
                else:
                    messagebox.showerror("Conexion",
                        f"No se pudo abrir {puerto}.")
                    self.cx=None
            elif tipo=="wifi":
                ip=self.e_ip.get().strip()
                port=int(self.e_wport.get().strip())
                self.cx=ConexionRadwag(on_dato=self._on_dato)
                self.cx.conectar(ip,port)
                self.btn_cx.config(text="Desconectar",bg=RED)
                self.lbl_cx.config(text=f"WiFi {ip}:{port}",fg=GREEN)

    def _on_dato(self,valor,raw):
        self.after(0,self._procesar,valor,raw)

    def _procesar(self,valor,raw):
        if raw=="__CONNECTED__":
            self.lbl_cx.config(text="WiFi conectado",fg=GREEN); return
        if raw=="__DISCONNECTED__":
            self.lbl_cx.config(text="Reconectando...",fg=YELLOW); return
        if valor is None: return
        if self.tab_activa:
            self.panels_aba[self.tab_activa].recibir(valor)

    def _registrar_manual(self):
        if not self.tab_activa: return
        panel=self.panels_aba[self.tab_activa]
        if panel.paso==0:
            messagebox.showwarning("Sin ciclo",
                "Presiona Iniciar nueva caracterizacion primero.")
            return
        txt=self.e_manual.get().strip().replace(",",".")
        try: valor=float(txt)
        except ValueError:
            messagebox.showwarning("Valor invalido",
                "Ingresa un numero valido (p.ej. 10000,4)")
            return
        self.e_manual.delete(0,"end"); self.e_manual.focus()
        panel.recibir(valor)


class App:
    def __init__(self,root):
        self.root=root
        self.root.title(
            "METROMECANICA — Caracterizacion de Balanzas"
            " | NMP 004:2007 | ISO/IEC 17025")
        self.root.geometry("1400x860")
        self.root.configure(bg=BG)
        self.root.minsize(1100,700)
        self.root.protocol("WM_DELETE_WINDOW",self._cerrar)
        self.hist=cargar_hist()
        self.panel_balanza_activo=None
        self.panel_aba_completado=None
        self._build_ui(); self._tick()

    def _cerrar(self):
        for pb in self.balanzas.values():
            if pb.cx and getattr(pb.cx,"activo",False):
                pb.cx.desconectar()
        self.root.destroy()

    def _build_ui(self):
        tk.Frame(self.root,bg=ACCENT,height=3).pack(fill="x")
        hdr=tk.Frame(self.root,bg=BG,padx=20,pady=8); hdr.pack(fill="x")
        tk.Label(hdr,text="METROMECANICA",bg=BG,fg=ACCENT,
                 font=("Georgia",14,"bold")).pack(side="left")
        tk.Label(hdr,
            text="  Caracterizacion de Balanzas  |  "
                 "NMP 004:2007  |  Clase M2  |  ISO/IEC 17025",
            bg=BG,fg=TXT_DIM,
            font=("Georgia",8,"italic")).pack(side="left")
        self.lbl_reloj=tk.Label(hdr,bg=BG,fg=TXT_DIM,
            font=("Courier New",9)); self.lbl_reloj.pack(side="right")
        tk.Frame(self.root,bg=BORDER,height=1).pack(fill="x")

        foot=tk.Frame(self.root,bg=PANEL); foot.pack(fill="x",side="bottom")
        tk.Frame(foot,bg=BORDER,height=1).pack(fill="x")
        f2=tk.Frame(foot,bg=PANEL,padx=12,pady=6); f2.pack(fill="x")
        self.btn_guardar=tk.Button(f2,text="GUARDAR CICLO",
            bg=GREEN,fg="white",font=("Georgia",10,"bold"),
            relief="flat",padx=16,pady=5,state="disabled",
            command=self._guardar)
        self.btn_guardar.pack(side="left",padx=(0,10))
        tk.Button(f2,text="EXPORTAR EXCEL",
            bg=ACCENT2,fg="white",font=FN_BOLD,
            relief="flat",padx=12,pady=5,
            command=self._exportar).pack(side="left",padx=(0,6))
        tk.Button(f2,text="BORRAR HISTORIAL",
            bg="#7f1d1d",fg="white",font=FN_UI,
            relief="flat",padx=10,pady=5,
            command=self._borrar).pack(side="left")
        tk.Label(f2,
            text="NMP 004:2007  |  Coma decimal INACAL  |  15 dec en s(DI)",
            bg=PANEL,fg=TXT_DIM,font=FN_SM).pack(side="right")

        self._build_balanzas()

    def _build_balanzas(self):
        cfgs={
            "radwag":{"nombre":"RADWAG AS 82/220.X2","tipo":"wifi",
                "ip":RADWAG_IP,"port":RADWAG_PORT,"color":TEAL,
                "pesas":PESAS_RADWAG,
                "tab_lbl":"  RADWAG AS 82/220  WiFi  0,0001 g  "},
            "biobase":{"nombre":"BIOBASE BL-5000","tipo":"serial",
                "puerto":"COM6","baud":9600,"color":ACCENT2,
                "pesas":PESAS_BIOBASE,
                "tab_lbl":"  BIOBASE BL-5000  RS-232  0,01 g  "},
            "want":{"nombre":"WANT GT-30000TR","tipo":"manual",
                "color":ORANGE,"pesas":PESAS_WANT,
                "tab_lbl":"  WANT GT-30000TR  Manual  0,1 g  "},
        }
        tab_bar=tk.Frame(self.root,bg=PANEL2,height=36)
        tab_bar.pack(fill="x"); tab_bar.pack_propagate(False)
        content=tk.Frame(self.root,bg=BG)
        content.pack(fill="both",expand=True)
        self.balanzas={}; self.bal_frames={}
        self.bal_btns={}; self.bal_activa=None

        for key,cfg in cfgs.items():
            frame=tk.Frame(content,bg=BG); self.bal_frames[key]=frame
            pb=PanelBalanza(frame,cfg,self.hist,
                            on_completado=self._on_completado)
            pb.pack(fill="both",expand=True); self.balanzas[key]=pb
            btn=tk.Button(tab_bar,text=cfg["tab_lbl"],
                bg=PANEL2,fg=TXT_DIM,font=("Georgia",9),
                relief="flat",padx=14,pady=6,
                command=lambda k=key:self._switch_bal(k))
            btn.pack(side="left"); self.bal_btns[key]=btn

        self._switch_bal("radwag")

    def _switch_bal(self,key):
        for k,frame in self.bal_frames.items():
            frame.pack_forget()
            self.bal_btns[k].config(bg=PANEL2,fg=TXT_DIM,font=("Georgia",9))
        color=self.balanzas[key].cfg["color"]
        self.bal_frames[key].pack(fill="both",expand=True)
        self.bal_btns[key].config(bg=color,fg="white",font=("Georgia",9,"bold"))
        self.bal_activa=key
        self.panel_balanza_activo=self.balanzas[key]
        self.btn_guardar.config(state="disabled",
            text="GUARDAR CICLO",bg=GREEN)

    def _on_completado(self,panel_aba):
        self.panel_aba_completado=panel_aba
        self.btn_guardar.config(state="normal",
            text="GUARDAR CICLO  <-- PRESIONA AQUI",
            bg="#15803d")

    def _guardar(self):
        if self.panel_aba_completado:
            self.panel_aba_completado.guardar_ciclo()
            self.panel_aba_completado=None
        self.btn_guardar.config(state="disabled",
            text="GUARDAR CICLO",bg=GREEN)

    def _tick(self):
        self.lbl_reloj.config(
            text=datetime.now().strftime("%Y-%m-%d  %H:%M:%S"))
        self.root.after(1000,self._tick)

    def _exportar(self):
        if not self.panel_balanza_activo: return
        pb=self.panel_balanza_activo
        panel=pb.panel_activo()
        if not panel:
            messagebox.showinfo("Sin panel","Selecciona una pesa."); return
        ciclos=self.hist.get(panel.key,{}).get("ciclos",[])
        if not ciclos:
            messagebox.showinfo("Sin datos","No hay ciclos para esta pesa."); return
        try:
            import openpyxl
            from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
        except ImportError:
            messagebox.showerror("openpyxl","pip install openpyxl"); return

        ts=datetime.now().strftime("%Y%m%d_%H%M%S")
        pesa_disp=panel.pesa[1]
        ruta=filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")],
            initialfile=f"carac_{panel.key.replace(' ','_')}_{ts}.xlsx")
        if not ruta: return

        def fill(h): return PatternFill("solid",fgColor=h)
        def font(bold=False,color="000000",sz=10):
            return Font(bold=bold,color=color,size=sz)
        def aln(h="center",v="center",wrap=False):
            return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
        def brd():
            s=Side(style="thin",color="AAAAAA")
            return Border(left=s,right=s,top=s,bottom=s)
        def brd_med():
            s=Side(style="medium",color="4472C4")
            return Border(left=s,right=s,top=s,bottom=s)

        wb=openpyxl.Workbook()
        sdis=[c["sdi"] for c in ciclos]
        ords={1:"1er",2:"2do",3:"3er",4:"4to",5:"5to",
              6:"6to",7:"7mo",8:"8vo",9:"9no",10:"10mo"}

        for idx_c,ciclo in enumerate(ciclos):
            n_c=ciclo["n"]
            ws=wb.active if idx_c==0 else wb.create_sheet(f"Ciclo {n_c}")
            ws.title=f"Ciclo {n_c}"
            for col,w in zip("ABCDE",[6,14,14,14,14]):
                ws.column_dimensions[col].width=w
            ws.column_dimensions["G"].width=44

            ws.merge_cells("A1:E1")
            c=ws["A1"]
            c.value=(f"METROMECANICA  {pb.cfg['nombre']}"
                     f"  |  NMP 004:2007  |  ISO/IEC 17025")
            c.font=font(True,"FFFFFF",10); c.fill=fill("17375E")
            c.alignment=aln(); c.border=brd(); ws.row_dimensions[1].height=16

            ws.merge_cells("A2:E2"); c=ws["A2"]
            c.value=(f"Pesa: {pesa_disp}     "
                     f"ID: {ciclo.get('pat_id','---')}     "
                     f"Cert.: {ciclo.get('pat_cert','---')}     "
                     f"Fecha Cal.: {ciclo.get('pat_fecha','---')}")
            c.font=font(color="404040",sz=9); c.fill=fill("F2F2F2")
            c.alignment=aln("left"); c.border=brd(); ws.row_dimensions[2].height=14

            ws.merge_cells("A3:E3"); c=ws["A3"]
            c.value=(f"Caracterizacion de la Balanza  "
                     f"Pesa {pesa_disp}  "
                     f"{ords.get(n_c,str(n_c))} Ciclo")
            c.font=font(True,sz=11); c.fill=fill("BDD7EE")
            c.alignment=aln(); c.border=brd(); ws.row_dimensions[3].height=18

            hdrs=[("N","A"),("PATRON (Ir1)\ng","B"),
                  ("CALIBRAR (It)\ng","C"),
                  ("PATRON (Ir2)\ng","D"),("DI\ng","E")]
            for txt,col in hdrs:
                cel=ws[f"{col}4"]; cel.value=txt
                cel.font=font(True,sz=10); cel.fill=fill("BDD7EE")
                cel.alignment=aln("center",wrap=True)
                cel.border=brd()
            ws.row_dimensions[4].height=28

            dd=panel.dd; lects=ciclo.get("lecturas",[])
            for j,l in enumerate(lects,1):
                fila=4+j
                ir1=l["ir1"]; it=l["it"]; ir2=l["ir2"]
                di=delta_i(ir1,it,ir2)
                for col,v in zip(["A","B","C","D","E"],[j,ir1,it,ir2,di]):
                    cel=ws[f"{col}{fila}"]
                    cel.value=round(v,dd) if col!="A" else v
                    cel.font=font(sz=10)
                    cel.alignment=aln("center"); cel.border=brd()
                    if col in ("B","C","D","E"):
                        cel.number_format="0."+"0"*dd
                    if col=="E": cel.fill=fill("E2EFDA")
                ws.row_dimensions[fila].height=15

            fila_s=4+len(lects)+2
            ws.merge_cells(f"A{fila_s}:D{fila_s}")
            c=ws[f"A{fila_s}"]
            c.value=("s(DI) :  Desviacion estandar de las diferencias de "
                     "Lectura de la pesa a calibrar y la pesa de referencia")
            c.font=font(sz=9,color="404040"); c.alignment=aln("right")
            ws.row_dimensions[fila_s].height=28
            c_sdi=ws[f"E{fila_s}"]
            c_sdi.value=ciclo["sdi"]; c_sdi.font=font(True,sz=11)
            c_sdi.fill=fill("00B0F0"); c_sdi.alignment=aln("center")
            c_sdi.border=brd_med(); c_sdi.number_format="0.00000"

            fila_v=fila_s+1
            ws.merge_cells(f"A{fila_v}:C{fila_v}")
            c=ws[f"A{fila_v}"]; c.value="Varianza de DI"
            c.font=font(True,sz=10); c.fill=fill("D9EAD3")
            c.alignment=aln("right"); c.border=brd()
            vac=varianza_ciclos(sdis[:n_c]) if n_c>=2 else None
            c_var=ws[f"E{fila_v}"]
            if vac is not None:
                c_var.value=vac; c_var.font=font(True,sz=11,color="FFFFFF")
                c_var.fill=fill("0D9488"); c_var.number_format="0.0000000"
            else:
                c_var.value="min. 2 ciclos"
                c_var.font=font(sz=9,color="808080"); c_var.fill=fill("EEEEEE")
            c_var.alignment=aln("center"); c_var.border=brd_med()
            ws[f"D{fila_v}"].value="Para 2 o mas caracterizaciones"
            ws[f"D{fila_v}"].font=font(sz=8,color="606060")
            ws[f"D{fila_v}"].alignment=aln("left")
            ws.row_dimensions[fila_v].height=20

        if len(ciclos)>1:
            ws_r=wb.create_sheet("Resumen")
            for col,w in zip("ABCDE",[8,20,22,22,22]):
                ws_r.column_dimensions[col].width=w
            ws_r.merge_cells("A1:E1"); c=ws_r["A1"]
            c.value=f"RESUMEN  {pb.cfg['nombre']}  Pesa {pesa_disp}"
            c.font=font(True,"FFFFFF",11); c.fill=fill("17375E")
            c.alignment=aln(); ws_r.row_dimensions[1].height=18
            for txt,col in [("Ciclo","A"),("Fecha","B"),
                            ("ID Patron","C"),("s(DI) g","D"),("Var g2","E")]:
                cel=ws_r[f"{col}2"]; cel.value=txt
                cel.font=font(True); cel.fill=fill("BDD7EE")
                cel.border=brd(); cel.alignment=aln()
            ws_r.row_dimensions[2].height=16
            for i,c in enumerate(ciclos,1):
                fila=2+i
                vac=varianza_ciclos(sdis[:i]) if i>=2 else None
                for col,v in zip(["A","B","C","D","E"],
                        [c["n"],c["fecha"],c.get("pat_id","---"),
                         c["sdi"],vac if vac is not None else "---"]):
                    cel=ws_r[f"{col}{fila}"]
                    cel.value=v; cel.font=font(sz=10)
                    cel.alignment=aln("center"); cel.border=brd()
                    if col=="D" and isinstance(v,float): cel.number_format="0.00000"
                    if col=="E" and isinstance(v,float): cel.number_format="0.0000000"
                ws_r.row_dimensions[fila].height=14

        wb.save(ruta)
        messagebox.showinfo("Exportado",f"Excel guardado:\n{ruta}")
        if messagebox.askyesno("Abrir","Abrir el archivo ahora?"):
            abrir_archivo(ruta)

    def _borrar(self):
        if not self.panel_balanza_activo: return
        panel=self.panel_balanza_activo.panel_activo()
        if not panel: return
        if not self.hist.get(panel.key,{}).get("ciclos"):
            messagebox.showinfo("Sin datos","No hay historial para esta pesa."); return
        if messagebox.askyesno("Borrar",f"Eliminar historial de {panel.key}?"):
            self.hist[panel.key]={"ciclos":[]}
            guardar_hist(self.hist); panel.refresh()


if __name__=="__main__":
    root=tk.Tk()
    App(root)
    root.mainloop()