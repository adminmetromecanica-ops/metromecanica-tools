"""
╔══════════════════════════════════════════════════════════════╗
  METROMECANICA — Sistema de Caracterización de Balanzas
  NMP 004:2007 | Clase M2 | ISO/IEC 17025
  v5.1 — Correcciones + Condiciones ambientales
  - Fix: contador ciclos no se actualizaba tras guardar
  - Fix: exportar Excel/PDF desde botones globales y detalle
  - Fix: EMP NMP 004:2007 Clase M2 verificado
  - Nuevo: Condiciones ambientales inicio/fin por caracterizacion
    (Temperatura, Humedad relativa, Presion atmosferica)
  - Nuevo: Condiciones aparecen en Excel y PDF exportados
╚══════════════════════════════════════════════════════════════╝
  pip install pyserial matplotlib openpyxl reportlab pyttsx3
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import serial, serial.tools.list_ports
import socket
import json, os, re, math, threading, time, sys, subprocess
from datetime import datetime

try:
    import pyttsx3; VOZ_OK = True
except: VOZ_OK = False

import matplotlib
matplotlib.use("TkAgg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

# ══════════════════════════════════════════════════════════════
#  CONFIGURACIÓN Y DATOS
# ══════════════════════════════════════════════════════════════
DIR_APP   = os.path.dirname(os.path.abspath(__file__))
FILE_HIST = os.path.join(DIR_APP, "historial_v5.json")

N_LECTURAS = 10   # lecturas ABA por ciclo
N_CICLOS   = 10   # ciclos para caracterización completa
RADWAG_IP   = "192.168.18.65"
RADWAG_PORT = 4001

# (nominal_g, etiqueta, EMP_mg, decimales)
PESAS_RADWAG = [
    (1.0,   "1 g",    1.5,   4),
    (2.0,   "2 g",    3.0,   4),
    (5.0,   "5 g",    5.0,   4),
    (10.0,  "10 g",   10.0,  4),
    (20.0,  "20 g",   25.0,  4),
    (50.0,  "50 g",   50.0,  4),
    (100.0, "100 g",  100.0, 4),
    (200.0, "200 g",  150.0, 4),
]
PESAS_BIOBASE = [
    (1000.0, "1 kg",  500.0,  2),
    (2000.0, "2 kg",  1000.0, 2),
    (5000.0, "5 kg",  2500.0, 2),
]
PESAS_WANT = [
    (10000.0, "10 kg", 5000.0,  1),
    (15000.0, "15 kg", 7500.0,  1),
    (20000.0, "20 kg", 10000.0, 1),
    (25000.0, "25 kg", 12500.0, 1),
]

BALANZAS_CFG = {
    "RADWAG": {
        "nombre": "RADWAG AS 82/220.X2",
        "id_bal": "BAL-03",
        "tipo": "wifi",
        "ip": RADWAG_IP, "port": RADWAG_PORT,
        "resolucion": "0,0001 g",
        "pesas": PESAS_RADWAG,
        "color_hex": "#0d9488",
        "color_dark": "#0a7470",
        "icono": "⚖",
    },
    "BIOBASE": {
        "nombre": "BIOBASE BL-5000",
        "id_bal": "BAL-02",
        "tipo": "serial",
        "puerto": "COM6", "baud": 9600,
        "resolucion": "0,01 g",
        "pesas": PESAS_BIOBASE,
        "color_hex": "#2563eb",
        "color_dark": "#1d4ed8",
        "icono": "⚖",
    },
    "WANT": {
        "nombre": "WANT GT-30000TR",
        "id_bal": "BAL-01",
        "tipo": "manual",
        "resolucion": "0,1 g",
        "pesas": PESAS_WANT,
        "color_hex": "#ea580c",
        "color_dark": "#c2410c",
        "icono": "⚖",
    },
}

# ══════════════════════════════════════════════════════════════
#  PALETA — Diseño oscuro premium
# ══════════════════════════════════════════════════════════════
C = {
    "bg":       "#0a0f1e",
    "surface":  "#111827",
    "surface2": "#1a2235",
    "surface3": "#1f2d42",
    "border":   "#243044",
    "border2":  "#2e3c52",
    "text":     "#e2e8f0",
    "text2":    "#94a3b8",
    "text3":    "#4a5568",
    "accent":   "#38bdf8",
    "green":    "#10b981",
    "red":      "#f43f5e",
    "yellow":   "#fbbf24",
    "purple":   "#8b5cf6",
}

F = {
    "title":  ("Georgia", 22, "bold"),
    "h1":     ("Georgia", 13, "bold"),
    "h2":     ("Georgia", 11, "bold"),
    "h3":     ("Georgia", 9, "bold"),
    "body":   ("Georgia", 9),
    "small":  ("Georgia", 8),
    "mono":   ("Courier New", 10),
    "mono_l": ("Courier New", 13, "bold"),
    "mono_xl":("Courier New", 20, "bold"),
    "mono_sm":("Courier New", 8),
}

ORDINAL = {1:"1er",2:"2do",3:"3er",4:"4to",5:"5to",
           6:"6to",7:"7mo",8:"8vo",9:"9no",10:"10mo"}

# ── Sesiones de caracterización ───────────────────────────────
MAX_SESIONES_ANIO = 2   # máximo de caracterizaciones por año
# Estructura historial:
# hist[key] = {"sesiones": [
#   {"id_sesion": "2026-S1", "anio": 2026, "num": 1,
#    "fecha_inicio": "...", "ciclos": [...], "alertas": []},
#   ...
# ]}

def sesion_actual(hist, key):
    """Retorna la sesión activa (incompleta) o None."""
    sesiones = hist.get(key, {}).get("sesiones", [])
    for s in sesiones:
        if len(s.get("ciclos", [])) < N_CICLOS:
            return s
    return None

def sesiones_anio(hist, key, anio=None):
    """Sesiones de un año específico."""
    if anio is None: anio = datetime.now().year
    sesiones = hist.get(key, {}).get("sesiones", [])
    return [s for s in sesiones if s.get("anio") == anio]

def nueva_sesion(hist, key):
    """Crea una nueva sesión de caracterización."""
    anio = datetime.now().year
    sas  = sesiones_anio(hist, key, anio)
    num  = len(sas) + 1
    if key not in hist:
        hist[key] = {"sesiones": []}
    sesion = {
        "id_sesion":    f"{anio}-S{num}",
        "anio":         anio,
        "num":          num,
        "fecha_inicio": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "ciclos":       [],
        "alertas":      [],
    }
    hist[key]["sesiones"].append(sesion)
    return sesion

def todos_sdis_historicos(hist, key):
    """Todos los s(ΔI) de todas las sesiones completadas."""
    sdis = []
    for s in hist.get(key, {}).get("sesiones", []):
        if len(s.get("ciclos", [])) >= N_CICLOS:
            sdis.extend([c["sdi"] for c in s["ciclos"]])
    return sdis


# ══════════════════════════════════════════════════════════════
#  SISTEMA DE ALERTAS DE DERIVA
# ══════════════════════════════════════════════════════════════
ALERTA_NINGUNA  = 0
ALERTA_INFO     = 1   # informativo (verde)
ALERTA_MODERADA = 2   # atención (amarillo)
ALERTA_CRITICA  = 3   # acción requerida (rojo)

def evaluar_deriva(sdi_nuevo, sdis_historicos, emp_mg):
    """
    Evalúa tres criterios de deriva. Retorna lista de alertas.
    Cada alerta: {"nivel": int, "criterio": str, "mensaje": str,
                  "recomendacion": str}
    """
    alertas = []
    if not sdis_historicos or len(sdis_historicos) < 2:
        return alertas

    media   = sum(sdis_historicos) / len(sdis_historicos)
    var     = varianza(sdis_historicos)
    sigma   = math.sqrt(var) if var else 0
    emp_g   = emp_mg / 1000.0
    lim_emp = emp_g * 0.5   # 50% del EMP

    # ── Criterio 1: Regla 2σ (desviación estadística) ─────────
    desviacion = abs(sdi_nuevo - media)
    if sigma > 0:
        z = desviacion / sigma
        if z > 3:
            alertas.append({
                "nivel":     ALERTA_CRITICA,
                "criterio":  "Deriva estadistica (3σ)",
                "valor":     f"z = {z:.2f}σ  (umbral: 3σ)",
                "mensaje":   f"El s(ΔI) se aleja {z:.1f} desviaciones de la media histórica.",
                "recomendacion": (
                    "ACCION REQUERIDA:\n"
                    "1. Suspender uso de la balanza para calibraciones.\n"
                    "2. Verificar limpieza y condiciones de las pesas patron.\n"
                    "3. Revisar nivel y estabilidad de la balanza.\n"
                    "4. Realizar nueva caracterizacion completa.\n"
                    "5. Si persiste, enviar balanza a mantenimiento."
                ),
            })
        elif z > 2:
            alertas.append({
                "nivel":     ALERTA_MODERADA,
                "criterio":  "Deriva estadistica (2σ)",
                "valor":     f"z = {z:.2f}σ  (umbral: 2σ)",
                "mensaje":   f"El s(ΔI) supera 2σ de la media histórica. Tendencia de deriva.",
                "recomendacion": (
                    "ATENCION:\n"
                    "1. Monitorear los próximos ciclos con atención.\n"
                    "2. Verificar condiciones ambientales (T, HR, P).\n"
                    "3. Revisar limpieza de plato y pesas.\n"
                    "4. Considerar adelantar la próxima caracterizacion."
                ),
            })

    # ── Criterio 2: Límite 50% EMP ─────────────────────────────
    if sdi_nuevo > lim_emp:
        alertas.append({
            "nivel":     ALERTA_CRITICA,
            "criterio":  "Limite EMP (50%)",
            "valor":     f"s(ΔI) = {fmt15(sdi_nuevo)} g  >  50% EMP = {fmt15(lim_emp)} g",
            "mensaje":   f"El s(ΔI) supera el 50% del EMP ({fmt_emp(emp_mg)}). "
                         f"La balanza puede estar fuera de especificación.",
            "recomendacion": (
                "ACCION REQUERIDA:\n"
                "1. Verificar que la pesa patron esté dentro de su EMP.\n"
                "2. Revisar ajuste y nivelación de la balanza.\n"
                "3. NO usar esta balanza para emitir certificados hasta resolver.\n"
                "4. Contactar al responsable técnico del laboratorio."
            ),
        })
    elif sdi_nuevo > lim_emp * 0.7:
        alertas.append({
            "nivel":     ALERTA_MODERADA,
            "criterio":  "Aproximacion al limite EMP",
            "valor":     f"s(ΔI) = {fmt15(sdi_nuevo)} g  (70% del umbral EMP)",
            "mensaje":   f"El s(ΔI) se acerca al 50% del EMP. Vigilancia recomendada.",
            "recomendacion": (
                "PRECAUCION:\n"
                "1. Aumentar frecuencia de monitoreo.\n"
                "2. Documentar condiciones ambientales con detalle.\n"
                "3. Revisar historial de mantenimiento de la balanza."
            ),
        })

    # ── Criterio 3: Varianza entre ciclos ─────────────────────
    if var is not None and sigma > 0:
        cv = sigma / media * 100  # coeficiente de variación %
        if cv > 20:
            alertas.append({
                "nivel":     ALERTA_CRITICA,
                "criterio":  "Varianza elevada (CV > 20%)",
                "valor":     f"CV = {cv:.1f}%  (umbral: 20%)",
                "mensaje":   f"Alta dispersión entre ciclos (CV={cv:.1f}%). "
                             f"Resultados poco reproducibles.",
                "recomendacion": (
                    "ACCION REQUERIDA:\n"
                    "1. Revisar protocolo de medición ABA.\n"
                    "2. Verificar estabilidad de indicación de la balanza.\n"
                    "3. Controlar vibraciones y corrientes de aire.\n"
                    "4. Repetir la caracterizacion en condiciones controladas."
                ),
            })
        elif cv > 10:
            alertas.append({
                "nivel":     ALERTA_MODERADA,
                "criterio":  "Varianza moderada (CV > 10%)",
                "valor":     f"CV = {cv:.1f}%  (umbral: 10%)",
                "mensaje":   f"Dispersión moderada entre ciclos (CV={cv:.1f}%).",
                "recomendacion": (
                    "ATENCION:\n"
                    "1. Verificar condiciones de medición (temperatura estable).\n"
                    "2. Asegurar tiempo de estabilización de las pesas.\n"
                    "3. Revisar si hay corrientes de aire o vibraciones."
                ),
            })

    return alertas

def nivel_max(alertas):
    """Nivel máximo de alerta en una lista."""
    if not alertas: return ALERTA_NINGUNA
    return max(a["nivel"] for a in alertas)

COLOR_ALERTA = {
    ALERTA_NINGUNA:  C["green"] if C else "#10b981",
    ALERTA_INFO:     "#10b981",
    ALERTA_MODERADA: "#fbbf24",
    ALERTA_CRITICA:  "#f43f5e",
}
ICONO_ALERTA = {
    ALERTA_NINGUNA:  "✔",
    ALERTA_INFO:     "ℹ",
    ALERTA_MODERADA: "⚠",
    ALERTA_CRITICA:  "🚨",
}


# ══════════════════════════════════════════════════════════════
#  UTILIDADES
# ══════════════════════════════════════════════════════════════
def fmt(v, d):
    return format(v, f".{d}f").replace(".", ",")

def fmt15(v):
    return format(v, ".15f").replace(".", ",")

def fmt_emp(mg):
    return f"{mg/1000:g} g" if mg >= 1000 else f"{mg:g} mg"

def delta_i(ir1, it, ir2):
    return it - (ir1 + ir2) / 2.0

def s_delta(lects):
    if len(lects) < 2: return 0.0
    d = [delta_i(l["ir1"],l["it"],l["ir2"]) for l in lects]
    n = len(d); m = sum(d)/n
    return math.sqrt(sum((x-m)**2 for x in d)/(n-1))

def varianza(sdis):
    n = len(sdis)
    if n < 2: return None
    m = sum(sdis)/n
    return sum((s-m)**2 for s in sdis)/(n-1)

def parsear_serial(raw):
    m = re.search(r'([+-]?\s*\d+\.?\d*)\s*(g|kg)', raw, re.I)
    if m:
        try:
            v = float(m.group(1).replace(" ",""))
            return v*1000 if m.group(2).lower()=="kg" else v
        except: pass
    return None

def parsear_radwag(raw):
    m = re.search(r'([+-]?\s*\d+\.?\d*)\s*g', raw, re.I)
    if m:
        try: return float(m.group(1).replace(" ",""))
        except: pass
    return None

def cargar(): 
    try:
        with open(FILE_HIST,"r",encoding="utf-8") as f: return json.load(f)
    except: return {}

def guardar(h):
    with open(FILE_HIST,"w",encoding="utf-8") as f:
        json.dump(h,f,indent=2,ensure_ascii=False)

def hablar(t):
    if not VOZ_OK: return
    def _r():
        try:
            e=pyttsx3.init(); e.setProperty("rate",150)
            for v in e.getProperty("voices"):
                if "spanish" in v.name.lower(): e.setProperty("voice",v.id); break
            e.say(t); e.runAndWait(); e.stop()
        except: pass
    threading.Thread(target=_r,daemon=True).start()

def abrir(ruta):
    if sys.platform=="win32": os.startfile(ruta)
    elif sys.platform=="darwin": subprocess.call(["open",ruta])
    else: subprocess.call(["xdg-open",ruta])


# ══════════════════════════════════════════════════════════════
#  CONEXIONES
# ══════════════════════════════════════════════════════════════
class ConexionSerial:
    def __init__(self,cb): self.cb=cb; self.ser=None; self.activo=False
    def conectar(self,puerto,baud=9600):
        try:
            self.ser=serial.Serial(port=puerto,baudrate=baud,
                bytesize=serial.EIGHTBITS,parity=serial.PARITY_NONE,
                stopbits=serial.STOPBITS_ONE,timeout=3)
            self.activo=True
            threading.Thread(target=self._loop,daemon=True).start()
            return True
        except: return False
    def desconectar(self):
        self.activo=False
        if self.ser:
            try: self.ser.close()
            except: pass
    def _loop(self):
        while self.activo:
            try:
                if self.ser and self.ser.in_waiting>0:
                    raw=self.ser.readline().decode("ascii",errors="ignore").strip()
                    if raw:
                        v=parsear_serial(raw)
                        if v is not None: self.cb(v,raw)
                time.sleep(0.05)
            except: break

class ConexionRadwag:
    def __init__(self,cb): self.cb=cb; self.sock=None; self.activo=False
    def conectar(self,ip,port):
        self.ip=ip; self.port=port; self.activo=True
        threading.Thread(target=self._loop,daemon=True).start()
        return True
    def desconectar(self):
        self.activo=False
        if self.sock:
            try: self.sock.close()
            except: pass; self.sock=None
    def _conectar(self):
        if self.sock:
            try: self.sock.close()
            except: pass; self.sock=None
        try:
            s=socket.socket(socket.AF_INET,socket.SOCK_STREAM)
            s.setsockopt(socket.SOL_SOCKET,socket.SO_REUSEADDR,1)
            s.settimeout(5); s.connect((self.ip,self.port))
            s.settimeout(2); self.sock=s; return True
        except: return False
    def _loop(self):
        while self.activo:
            if not self._conectar(): time.sleep(3); continue
            self.cb(None,"__CONNECTED__"); buf=""
            while self.activo:
                try:
                    data=self.sock.recv(256)
                    if not data: break
                    buf+=data.decode("ascii",errors="ignore")
                    while "\r\n" in buf:
                        ln,buf=buf.split("\r\n",1); ln=ln.strip()
                        if ln:
                            v=parsear_radwag(ln)
                            if v is not None: self.cb(v,ln)
                except socket.timeout: continue
                except: break
            if self.activo: self.cb(None,"__DISCONNECTED__"); time.sleep(2)


# ══════════════════════════════════════════════════════════════
#  WIDGET AUXILIAR: SEPARATOR
# ══════════════════════════════════════════════════════════════
def sep(parent, color=None, h=1, pady=4):
    tk.Frame(parent, bg=color or C["border"], height=h).pack(
        fill="x", pady=pady)


# ══════════════════════════════════════════════════════════════
#  VENTANA DE CARACTERIZACIÓN — ciclo ABA completo
# ══════════════════════════════════════════════════════════════
class VentanaCarac(tk.Toplevel):
    """
    Ventana modal para realizar un ciclo ABA de 10 lecturas
    para una pesa específica de una balanza específica.
    """
    def __init__(self, parent, balanza_key, pesa, hist, cx_getter,
                 on_guardado):
        super().__init__(parent)
        self.bkey      = balanza_key
        self.bcfg      = BALANZAS_CFG[balanza_key]
        self.pesa      = pesa          # (nominal, label, EMP, decimales)
        self.hist      = hist
        self.cx_getter = cx_getter
        self.on_guardado = on_guardado

        self.paso       = 0
        self.lecturas   = []
        self.tmp_ir1    = None
        self.tmp_it     = None
        self.ultimo_sdi = None

        color = self.bcfg["color_hex"]
        self.title(f"Ciclo ABA — {self.bcfg['nombre']} — {pesa[1]}")
        self.geometry("820x700")
        self.configure(bg=C["bg"])
        self.resizable(True, True)
        self.grab_set()

        self._build()
        self._iniciar_auto()

    @property
    def key(self): return f"{self.bcfg['nombre']}_{self.pesa[1]}"

    @property
    def dd(self): return self.pesa[3]

    def _iniciar_auto(self):
        """Iniciar ciclo automáticamente si hay conexión."""
        if self.bcfg["tipo"] == "manual":
            self._iniciar()
        else:
            cx = self.cx_getter(self.bkey)
            if cx and getattr(cx,"activo",False):
                self._iniciar()

    def _build(self):
        color = self.bcfg["color_hex"]

        # ── Header ─────────────────────────────────────────
        hdr = tk.Frame(self, bg=color, padx=20, pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text=f"  {self.bcfg['icono']}  CICLO ABA",
                 bg=color, fg="white",
                 font=("Georgia",12,"bold")).pack(side="left")
        tk.Label(hdr,
                 text=f"{self.bcfg['id_bal']}  —  {self.bcfg['nombre']}  "
                      f"|  Pesa: {self.pesa[1]}",
                 bg=color, fg="white",
                 font=("Georgia",9)).pack(side="left", padx=16)
        # Sin "Ciclo #N/10" en el header

        body = tk.Frame(self, bg=C["bg"])
        body.pack(fill="both", expand=True, padx=16, pady=12)

        col_izq = tk.Frame(body, bg=C["bg"], width=340)
        col_izq.pack(side="left", fill="y", padx=(0,12))
        col_izq.pack_propagate(False)
        col_der = tk.Frame(body, bg=C["bg"])
        col_der.pack(side="right", fill="both", expand=True)

        self._build_izq(col_izq, color)
        self._build_der(col_der, color)

        # ── Footer ─────────────────────────────────────────
        foot = tk.Frame(self, bg=C["surface"], padx=16, pady=10)
        foot.pack(fill="x", side="bottom")
        self.btn_guardar = tk.Button(foot,
            text="  GUARDAR CICLO  ",
            bg=C["green"], fg="white",
            font=("Georgia",11,"bold"),
            relief="flat", padx=20, pady=8,
            state="disabled",
            command=self._guardar)
        self.btn_guardar.pack(side="left", padx=(0,10))
        tk.Button(foot,
            text="Cancelar",
            bg=C["surface3"], fg=C["text2"],
            font=F["body"], relief="flat", padx=14, pady=8,
            command=self.destroy).pack(side="left")
        self.lbl_estado_foot = tk.Label(foot,
            text="Esperando inicio...",
            bg=C["surface"], fg=C["text2"],
            font=F["small"])
        self.lbl_estado_foot.pack(side="right")

    def _build_izq(self, parent, color):
        # Datos patrón
        card = tk.Frame(parent, bg=C["surface"], padx=14, pady=12)
        card.pack(fill="x", pady=(0,10))
        tk.Label(card, text="DATOS PESA PATRÓN",
                 bg=C["surface"], fg=color,
                 font=("Georgia",7,"bold")).pack(anchor="w")
        tk.Frame(card, bg=C["border"], height=1).pack(
            fill="x", pady=(4,8))

        def campo(lbl, attr, val=""):
            r = tk.Frame(card, bg=C["surface"]); r.pack(fill="x", pady=2)
            tk.Label(r, text=lbl, bg=C["surface"], fg=C["text2"],
                     font=F["small"], width=13, anchor="w").pack(side="left")
            e = tk.Entry(r, font=("Courier New",9),
                         bg=C["surface3"], fg=C["text"],
                         insertbackground=color,
                         relief="flat", bd=0,
                         highlightthickness=1,
                         highlightbackground=C["border"],
                         highlightcolor=color,
                         width=16)
            e.insert(0, val); e.pack(side="left", padx=(4,0), ipady=4)
            setattr(self, attr, e)

        campo("ID / Código:", "e_id")
        campo("N° Certificado:", "e_cert")
        campo("Fecha Cal.:", "e_fecha",
              datetime.now().strftime("%Y-%m-%d"))

        # ── Condiciones ambientales ───────────────────────
        sep(parent)
        card_amb = tk.Frame(parent, bg=C["surface"], padx=14, pady=12)
        card_amb.pack(fill="x", pady=(0,6))
        tk.Label(card_amb, text="CONDICIONES AMBIENTALES — INICIO",
                 bg=C["surface"], fg=color,
                 font=("Georgia",7,"bold")).pack(anchor="w")
        tk.Frame(card_amb, bg=C["border"], height=1).pack(
            fill="x", pady=(4,8))

        def campo_amb(lbl, attr, unidad, val="", w=7):
            r=tk.Frame(card_amb, bg=C["surface"])
            r.pack(fill="x", pady=2)
            tk.Label(r, text=lbl, bg=C["surface"], fg=C["text2"],
                     font=F["small"], width=8, anchor="w").pack(side="left")
            e=tk.Entry(r, font=("Courier New",9),
                       bg=C["surface3"], fg=C["text"],
                       insertbackground=color,
                       relief="flat", bd=0,
                       highlightthickness=1,
                       highlightbackground=C["border"],
                       highlightcolor=color, width=w)
            e.insert(0, val); e.pack(side="left", padx=(4,0), ipady=4)
            tk.Label(r, text=unidad, bg=C["surface"], fg=C["text3"],
                     font=F["small"]).pack(side="left", padx=4)
            setattr(self, attr, e)

        campo_amb("Temp.:",    "e_temp_i",  "°C")
        campo_amb("HR:",       "e_hum_i",   "%")
        campo_amb("Presión:",  "e_pres_i",  "hPa", w=8)

        sep(parent)
        card_fin = tk.Frame(parent, bg=C["surface"], padx=14, pady=12)
        card_fin.pack(fill="x", pady=(0,6))
        tk.Label(card_fin, text="CONDICIONES AMBIENTALES — FIN",
                 bg=C["surface"], fg=color,
                 font=("Georgia",7,"bold")).pack(anchor="w")
        tk.Frame(card_fin, bg=C["border"], height=1).pack(
            fill="x", pady=(4,8))

        def campo_fin(lbl, attr, unidad, w=7):
            r=tk.Frame(card_fin, bg=C["surface"])
            r.pack(fill="x", pady=2)
            tk.Label(r, text=lbl, bg=C["surface"], fg=C["text2"],
                     font=F["small"], width=8, anchor="w").pack(side="left")
            e=tk.Entry(r, font=("Courier New",9),
                       bg=C["surface3"], fg=C["text"],
                       insertbackground=color,
                       relief="flat", bd=0,
                       highlightthickness=1,
                       highlightbackground=C["border"],
                       highlightcolor=color, width=w)
            e.pack(side="left", padx=(4,0), ipady=4)
            tk.Label(r, text=unidad, bg=C["surface"], fg=C["text3"],
                     font=F["small"]).pack(side="left", padx=4)
            setattr(self, attr, e)

        tk.Label(card_fin,
                 text="Completa al finalizar los 10 ciclos",
                 bg=C["surface"], fg=C["text3"],
                 font=("Georgia",7,"italic")).pack(anchor="w", pady=(0,4))
        campo_fin("Temp.:",   "e_temp_f",  "°C")
        campo_fin("HR:",      "e_hum_f",   "%")
        campo_fin("Presión:", "e_pres_f",  "hPa", w=8)

        # Ingreso manual (WANT)
        if self.bcfg["tipo"] == "manual":
            sep(parent)
            card_m = tk.Frame(parent, bg=C["surface"], padx=14, pady=12)
            card_m.pack(fill="x", pady=(0,10))
            tk.Label(card_m, text="INGRESO MANUAL",
                     bg=C["surface"], fg=color,
                     font=("Georgia",7,"bold")).pack(anchor="w")
            tk.Frame(card_m, bg=C["border"], height=1).pack(
                fill="x", pady=(4,8))
            tk.Label(card_m,
                     text="Escribe la lectura y presiona\nREGISTRAR o ENTER",
                     bg=C["surface"], fg=C["text2"],
                     font=F["small"], justify="left").pack(anchor="w")
            self.e_manual = tk.Entry(card_m,
                font=("Courier New", 26, "bold"),
                bg=C["bg"], fg=C["green"],
                insertbackground=color,
                relief="flat", bd=0,
                highlightthickness=2,
                highlightbackground=C["border"],
                highlightcolor=color,
                justify="right", width=12)
            self.e_manual.pack(fill="x", pady=(8,4), ipady=6)
            tk.Label(card_m, text="gramos",
                     bg=C["surface"], fg=C["text3"],
                     font=F["small"]).pack(anchor="e")
            tk.Button(card_m,
                text="  ✔  REGISTRAR  ",
                bg=color, fg="white",
                font=("Georgia",11,"bold"),
                relief="flat", pady=8,
                command=self._registrar_manual).pack(
                fill="x", pady=(6,0))
            self.e_manual.bind("<Return>",
                lambda e: self._registrar_manual())
            self.e_manual.focus()

        # Estado paso actual
        sep(parent)
        self.card_paso = tk.Frame(parent, bg=C["surface"],
                                   padx=14, pady=12)
        self.card_paso.pack(fill="x")
        tk.Label(self.card_paso, text="INSTRUCCIÓN ACTUAL",
                 bg=C["surface"], fg=C["text3"],
                 font=("Georgia",7,"bold")).pack(anchor="w")
        tk.Frame(self.card_paso, bg=C["border"], height=1).pack(
            fill="x", pady=(4,8))
        self.lbl_paso = tk.Label(self.card_paso,
            text="Iniciando...",
            bg=C["surface"], fg=C["yellow"],
            font=("Georgia",9,"bold"),
            wraplength=300, justify="left")
        self.lbl_paso.pack(anchor="w")

        # Parcial Ir1/It/Ir2
        sep(parent)
        card_p = tk.Frame(parent, bg=C["surface"], padx=14, pady=10)
        card_p.pack(fill="x")
        for txt, attr in [("Ir1","lbl_ir1"),("It","lbl_it"),("Ir2","lbl_ir2")]:
            r = tk.Frame(card_p, bg=C["surface"]); r.pack(fill="x", pady=2)
            tk.Label(r, text=txt, bg=C["surface"], fg=C["text3"],
                     font=F["small"], width=4, anchor="w").pack(side="left")
            lv = tk.Label(r, text="—",
                bg=C["surface3"], fg=C["text2"],
                font=("Courier New",9,"bold"),
                anchor="e", padx=8, pady=3,
                relief="flat", width=16)
            lv.pack(side="left", padx=4)
            tk.Label(r, text="g", bg=C["surface"],
                     fg=C["text3"], font=F["small"]).pack(side="left")
            setattr(self, attr, lv)
        self.lbl_di = tk.Label(card_p,
            text="ΔI  =  —",
            bg=C["surface"], fg=C["accent"],
            font=("Courier New",9,"bold"))
        self.lbl_di.pack(anchor="w", pady=(6,0))

    def _build_der(self, parent, color):
        # Contador
        top = tk.Frame(parent, bg=C["surface"], padx=16, pady=10)
        top.pack(fill="x", pady=(0,8))
        self.lbl_count = tk.Label(top,
            text=f"0 / {N_LECTURAS}",
            bg=C["surface"], fg=color,
            font=("Courier New",28,"bold"))
        self.lbl_count.pack(side="left")
        tk.Label(top, text="lecturas completadas",
                 bg=C["surface"], fg=C["text3"],
                 font=F["small"]).pack(side="left", padx=10, anchor="s")

        # s(ΔI) y varianza en tiempo real
        stats = tk.Frame(parent, bg=C["surface3"],
                         padx=12, pady=8)
        stats.pack(fill="x", pady=(0,8))
        for lbl, attr, w in [
            ("s(ΔI) parcial:", "lbl_sdi_live", 22),
            ("Var. intra-ciclo:", "lbl_var_live", 22)]:
            r = tk.Frame(stats, bg=C["surface3"])
            r.pack(fill="x", pady=1)
            tk.Label(r, text=lbl, bg=C["surface3"],
                     fg=C["text2"], font=F["small"],
                     width=16, anchor="w").pack(side="left")
            lv = tk.Label(r, text="—",
                bg=C["surface3"], fg=C["accent"],
                font=("Courier New",8,"bold"),
                anchor="w", width=w)
            lv.pack(side="left")
            tk.Label(r, text="g", bg=C["surface3"],
                     fg=C["text3"], font=F["small"]).pack(side="left")
            setattr(self, attr, lv)

        # Tabla de lecturas
        cols = ("N°","Ir1  (g)","It  (g)","Ir2  (g)","ΔI  (g)")
        self.tbl = ttk.Treeview(parent, columns=cols,
                                show="headings", height=12)
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview",
            background=C["surface"],
            foreground=C["text"],
            fieldbackground=C["surface"],
            rowheight=22,
            font=("Courier New",9))
        style.configure("Treeview.Heading",
            background=C["surface2"],
            foreground=C["text2"],
            font=("Georgia",8,"bold"),
            relief="flat")
        style.map("Treeview",
            background=[("selected", color)],
            foreground=[("selected","white")])
        for col, w in zip(cols, [30,110,110,110,100]):
            self.tbl.heading(col, text=col)
            self.tbl.column(col, width=w, anchor="center", minwidth=w)
        sb = ttk.Scrollbar(parent, orient="vertical",
                           command=self.tbl.yview)
        sb.pack(side="right", fill="y")
        self.tbl.configure(yscrollcommand=sb.set)
        self.tbl.pack(fill="both", expand=True)

        # s(ΔI) final destacado
        card_f = tk.Frame(parent, bg=C["surface"], padx=14, pady=10)
        card_f.pack(fill="x", pady=(8,0))
        r1 = tk.Frame(card_f, bg=C["surface"]); r1.pack(fill="x")
        tk.Label(r1, text="s(ΔI) final:",
                 bg=C["surface"], fg=C["text2"],
                 font=F["h3"]).pack(side="left")
        self.lbl_sdi_final = tk.Label(r1, text="—",
            bg=C["surface"], fg=C["green"],
            font=("Courier New",14,"bold"))
        self.lbl_sdi_final.pack(side="left", padx=10)
        tk.Label(r1, text="g", bg=C["surface"],
                 fg=C["text3"], font=F["body"]).pack(side="left")

    # ── Lógica ABA ─────────────────────────────────────────
    def _iniciar(self):
        self.paso = 1; self.lecturas = []
        self.tmp_ir1 = None; self.tmp_it = None
        self.ultimo_sdi = None
        for i in self.tbl.get_children(): self.tbl.delete(i)
        for a in ("lbl_ir1","lbl_it","lbl_ir2"):
            getattr(self,a).config(text="—", fg=C["text2"])
        self.lbl_di.config(text="ΔI  =  —", fg=C["accent"])
        self._upd_paso()

    def _upd_paso(self):
        n = len(self.lecturas)+1
        msgs = {
            1: f"Lect. {n}/{N_LECTURAS}  →  Ir1: Coloca PATRÓN y presiona PRINT",
            2: f"Lect. {n}/{N_LECTURAS}  →  It:  Coloca CALIBRAR y presiona PRINT",
            3: f"Lect. {n}/{N_LECTURAS}  →  Ir2: Coloca PATRÓN y presiona PRINT",
        }
        self.lbl_paso.config(
            text=msgs.get(self.paso,"—"),
            fg=C["yellow"] if self.paso>0 else C["text3"])
        self.lbl_estado_foot.config(
            text=f"Ciclo en progreso  |  Paso {self.paso}/3  |  "
                 f"Lectura {len(self.lecturas)+1}/{N_LECTURAS}")

    def recibir(self, valor):
        """Punto de entrada unificado para todos los tipos de conexión."""
        if self.paso == 0: return
        if self.paso == 1:
            self.tmp_ir1=valor; self.paso=2
            self.lbl_ir1.config(text=fmt(valor,self.dd)+" g",
                               fg=C["green"])
            self.lbl_it.config(text="—", fg=C["text2"])
            self.lbl_ir2.config(text="—", fg=C["text2"])
            self._upd_paso()
        elif self.paso == 2:
            self.tmp_it=valor; self.paso=3
            self.lbl_it.config(text=fmt(valor,self.dd)+" g",
                               fg=C["green"])
            self.lbl_ir2.config(text="esperando...", fg=C["yellow"])
            self._upd_paso()
        elif self.paso == 3:
            ir1=self.tmp_ir1; it=self.tmp_it; ir2=valor
            di=delta_i(ir1,it,ir2)
            self.lbl_ir2.config(text=fmt(ir2,self.dd)+" g",
                                fg=C["green"])
            self.lbl_di.config(
                text=f"ΔI  =  {fmt15(di)} g",
                fg=C["accent"])
            self.lecturas.append({"ir1":ir1,"it":it,"ir2":ir2})
            n=len(self.lecturas)
            self.tbl.insert("","end", values=(
                n,
                fmt(ir1,self.dd),
                fmt(it,self.dd),
                fmt(ir2,self.dd),
                fmt(di,min(self.dd,4))))
            hijos=self.tbl.get_children()
            if hijos: self.tbl.see(hijos[-1])
            self.lbl_count.config(text=f"{n} / {N_LECTURAS}")
            if n>=2:
                sdi_p=s_delta(self.lecturas)
                self.lbl_sdi_live.config(text=fmt15(sdi_p))
                deltas=[delta_i(l["ir1"],l["it"],l["ir2"])
                        for l in self.lecturas]
                mn=sum(deltas)/len(deltas)
                var=sum((d-mn)**2 for d in deltas)/(len(deltas)-1)
                self.lbl_var_live.config(text=fmt15(var))
            self.tmp_ir1=None; self.tmp_it=None
            for a in ("lbl_ir1","lbl_it","lbl_ir2"):
                getattr(self,a).config(text="—",fg=C["text2"])
            if n>=N_LECTURAS: self._completar()
            else: self.paso=1; self._upd_paso()

    def _registrar_manual(self):
        if self.paso==0: return
        txt=self.e_manual.get().strip().replace(",",".")
        try: valor=float(txt)
        except:
            messagebox.showwarning("Valor inválido",
                "Ingresa un número válido\n(ej: 10000,4)",
                parent=self); return
        self.e_manual.delete(0,"end")
        self.e_manual.focus()
        self.recibir(valor)

    def _completar(self):
        self.paso=0
        sdi=s_delta(self.lecturas); self.ultimo_sdi=sdi
        self.lbl_sdi_final.config(text=fmt15(sdi)+" g",
                                   fg=C["green"])
        n_sig=len(self.hist.get(self.key,{}).get("ciclos",[]))+1
        self.lbl_paso.config(
            text=f"✔  {ORDINAL.get(n_sig,str(n_sig))} ciclo completado\n"
                 f"Presiona GUARDAR CICLO",
            fg=C["green"])
        self.btn_guardar.config(state="normal")
        self.lbl_estado_foot.config(
            text=f"✔  Caracterizacion #{n_sig} lista para guardar",
            fg=C["green"])
        hablar(f"{ORDINAL.get(n_sig,str(n_sig))} ciclo completo")

    def _guardar(self):
        if self.ultimo_sdi is None: return
        ts   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        anio = datetime.now().year
        if self.key not in self.hist:
            self.hist[self.key] = {"sesiones":[]}
        if "sesiones" not in self.hist[self.key]:
            self.hist[self.key]["sesiones"] = []

        ses = sesion_actual(self.hist, self.key)
        if ses is None:
            sas   = sesiones_anio(self.hist, self.key, anio)
            n_sas = len(sas)
            if n_sas >= MAX_SESIONES_ANIO:
                resp = messagebox.askyesno(
                    "Limite anual",
                    f"Ya tienes {n_sas} caracterizaciones en {anio}.\n"
                    f"El limite recomendado es {MAX_SESIONES_ANIO}.\n\n"
                    f"Una caracterizacion adicional puede indicar\n"
                    f"que la balanza requiere atencion especial.\n\n"
                    f"Crear caracterizacion adicional?",
                    parent=self)
                if not resp: return
            ses = nueva_sesion(self.hist, self.key)

        ciclos = ses["ciclos"]
        n      = len(ciclos) + 1
        cond_i = {"temp":self.e_temp_i.get().strip(),
                  "hum": self.e_hum_i.get().strip(),
                  "pres":self.e_pres_i.get().strip()}
        cond_f = {"temp":self.e_temp_f.get().strip(),
                  "hum": self.e_hum_f.get().strip(),
                  "pres":self.e_pres_f.get().strip()}
        ciclos.append({
            "n":n,"fecha":ts,"sdi":self.ultimo_sdi,
            "pat_id":   self.e_id.get().strip(),
            "pat_cert": self.e_cert.get().strip(),
            "pat_fecha":self.e_fecha.get().strip(),
            "cond_inicio":cond_i,"cond_fin":cond_f,
            "lecturas":list(self.lecturas),
        })

        # Evaluacion de deriva
        sdis_hist   = todos_sdis_historicos(self.hist, self.key)
        sdis_sesion = [c["sdi"] for c in ciclos[:-1]]
        sdis_eval   = sdis_hist + sdis_sesion
        alertas     = evaluar_deriva(self.ultimo_sdi, sdis_eval, self.pesa[2])
        if alertas:
            if "alertas" not in ses: ses["alertas"] = []
            ses["alertas"].append({
                "fecha":ts,"ciclo":n,
                "sdi":self.ultimo_sdi,"items":alertas})

        guardar(self.hist)
        total = len(ciclos)
        hablar(f"Ciclo {n} guardado")
        if self.on_guardado: self.on_guardado()

        if alertas:
            self._mostrar_alertas(alertas, n)

        if total >= N_CICLOS:
            ses["fecha_fin"] = ts
            guardar(self.hist)
            hablar(f"Caracterizacion completa. Pesa {self.pesa[1]}.")
            sas_anio = sesiones_anio(self.hist, self.key, anio)
            messagebox.showinfo(
                "Caracterizacion Completa",
                f"Pesa: {self.pesa[1]}\n"
                f"Sesion: {ses['id_sesion']}\n"
                f"Ciclos ABA: {total}/{N_CICLOS}\n"
                f"s(DI) promedio: {fmt15(sum(c['sdi'] for c in ciclos)/len(ciclos))} g\n"
                f"Caracterizaciones en {anio}: {len(sas_anio)} / {MAX_SESIONES_ANIO}\n\n"
                f"Proxima caracterizacion recomendada: en 6 meses.",
                parent=self)
        self.destroy()

    def _mostrar_alertas(self, alertas, n_ciclo):
        nivel = nivel_max(alertas)
        ca_map = {ALERTA_MODERADA:"#fbbf24", ALERTA_CRITICA:"#f43f5e"}
        color_alerta = ca_map.get(nivel, "#10b981")
        win = tk.Toplevel(self)
        win.title(f"Alerta de Deriva — Ciclo {n_ciclo}")
        win.geometry("660x540")
        win.configure(bg=C["bg"])
        win.grab_set()
        icono = ICONO_ALERTA.get(nivel,"⚠")
        hdr = tk.Frame(win, bg=color_alerta, padx=20, pady=14)
        hdr.pack(fill="x")
        tk.Label(hdr,
            text=f"  {icono}  ALERTA DE DERIVA — Ciclo ABA {n_ciclo}",
            bg=color_alerta, fg="white",
            font=("Georgia",12,"bold")).pack(side="left")
        tk.Label(hdr,
            text=f"{self.pesa[1]}  |  s(DI)={fmt15(self.ultimo_sdi)} g",
            bg=color_alerta, fg="white",
            font=("Courier New",8)).pack(side="right")
        # Scroll
        sf = tk.Frame(win, bg=C["bg"]); sf.pack(fill="both",expand=True,padx=12,pady=8)
        cv = tk.Canvas(sf, bg=C["bg"], highlightthickness=0)
        sb2 = ttk.Scrollbar(sf, orient="vertical", command=cv.yview)
        inner = tk.Frame(cv, bg=C["bg"])
        cv.create_window((0,0), window=inner, anchor="nw")
        cv.configure(yscrollcommand=sb2.set)
        inner.bind("<Configure>", lambda e: cv.configure(scrollregion=cv.bbox("all")))
        sb2.pack(side="right", fill="y"); cv.pack(side="left", fill="both", expand=True)
        for a in alertas:
            ca = COLOR_ALERTA.get(a["nivel"], C["yellow"])
            ia = ICONO_ALERTA.get(a["nivel"],"⚠")
            card = tk.Frame(inner, bg=C["surface"],
                highlightthickness=1, highlightbackground=ca)
            card.pack(fill="x", pady=(0,8))
            ch = tk.Frame(card, bg=ca, padx=10, pady=6); ch.pack(fill="x")
            tk.Label(ch, text=f"{ia}  {a['criterio']}",
                bg=ca, fg="white",
                font=("Georgia",9,"bold")).pack(side="left")
            tk.Label(ch, text=a["valor"],
                bg=ca, fg="white",
                font=("Courier New",8)).pack(side="right")
            tk.Label(card, text=a["mensaje"],
                bg=C["surface"], fg=C["text"],
                font=("Georgia",9), wraplength=590,
                justify="left", padx=12, pady=6).pack(anchor="w")
            tk.Label(card, text=a["recomendacion"],
                bg=C["surface2"], fg=C["text2"],
                font=("Courier New",8), wraplength=570,
                justify="left", padx=12, pady=8).pack(fill="x",padx=8,pady=(0,8))
        foot = tk.Frame(win, bg=C["surface"], padx=16, pady=10)
        foot.pack(fill="x", side="bottom")
        tk.Label(foot,
            text="Alertas registradas en historial e incluidas en PDF/Excel.",
            bg=C["surface"], fg=C["text3"],
            font=("Georgia",7,"italic")).pack(side="left")
        tk.Button(foot, text="  Entendido  ",
            bg=color_alerta, fg="white",
            font=("Georgia",10,"bold"), relief="flat",
            padx=16, pady=6,
            command=win.destroy).pack(side="right")


# ══════════════════════════════════════════════════════════════
#  VENTANA DETALLE — historial + gráfico de una pesa
# ══════════════════════════════════════════════════════════════
class VentanaDetalle(tk.Toplevel):
    def __init__(self, parent, balanza_key, pesa, hist, on_borrar):
        super().__init__(parent)
        self.bkey    = balanza_key
        self.bcfg    = BALANZAS_CFG[balanza_key]
        self.pesa    = pesa
        self.hist    = hist
        self.on_borrar = on_borrar
        color = self.bcfg["color_hex"]
        self.title(f"Historial — {self.bcfg['id_bal']} — {self.bcfg['nombre']} — {pesa[1]}")
        self.geometry("900x620")
        self.configure(bg=C["bg"])
        self.resizable(True, True)
        self._build(color)

    @property
    def key(self): return f"{self.bcfg['nombre']}_{self.pesa[1]}"

    def _build(self, color):
        hdr = tk.Frame(self, bg=color, padx=20, pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text=f"HISTORIAL DE CARACTERIZACIONES",
                 bg=color, fg="white",
                 font=("Georgia",12,"bold")).pack(side="left")
        tk.Label(hdr,
                 text=f"{self.bcfg['id_bal']}  —  {self.bcfg['nombre']}  |  Pesa: {self.pesa[1]}",
                 bg=color, fg="white",
                 font=("Georgia",9)).pack(side="left", padx=16)
        tk.Button(hdr, text="✕ Borrar historial",
            bg=self.bcfg["color_dark"], fg="white",
            font=F["small"], relief="flat", padx=10, pady=4,
            command=self._borrar).pack(side="right")

        body = tk.Frame(self, bg=C["bg"])
        body.pack(fill="both", expand=True, padx=16, pady=12)

        col_izq = tk.Frame(body, bg=C["bg"])
        col_izq.pack(side="left", fill="both", expand=True,
                     padx=(0,10))
        col_der = tk.Frame(body, bg=C["bg"], width=340)
        col_der.pack(side="right", fill="y")
        col_der.pack_propagate(False)

        self._build_tabla(col_izq, color)
        self._build_grafico_y_stats(col_der, color)

    def _build_tabla(self, parent, color):
        tk.Label(parent, text="CICLOS REGISTRADOS",
                 bg=C["bg"], fg=color,
                 font=("Georgia",7,"bold")).pack(anchor="w")
        tk.Frame(parent, bg=C["border"], height=1).pack(
            fill="x", pady=(4,8))
        cols = ("N°","Fecha","ID Patrón","Cert.","s(ΔI) g","Var. g²")
        self.tbl = ttk.Treeview(parent, columns=cols,
                                show="headings", height=16)
        for col, w in zip(cols, [30,140,100,100,200,200]):
            self.tbl.heading(col, text=col)
            self.tbl.column(col, width=w, anchor="center", minwidth=30)
        sy = ttk.Scrollbar(parent, orient="vertical",
                           command=self.tbl.yview)
        sx = ttk.Scrollbar(parent, orient="horizontal",
                           command=self.tbl.xview)
        self.tbl.configure(yscrollcommand=sy.set,
                           xscrollcommand=sx.set)
        sy.pack(side="right", fill="y")
        self.tbl.pack(fill="both", expand=True)
        sx.pack(fill="x")
        self._poblar_tabla()

    def _poblar_tabla(self):
        for i in self.tbl.get_children(): self.tbl.delete(i)
        ciclos = self.hist.get(self.key,{}).get("ciclos",[])
        sdis = [c["sdi"] for c in ciclos]
        for i,c in enumerate(ciclos):
            var_s="—"
            if i>=1:
                v=varianza(sdis[:i+1])
                if v is not None: var_s=fmt15(v)
            self.tbl.insert("","end", values=(
                c["n"],c["fecha"],
                c.get("pat_id","—"),
                c.get("pat_cert","—"),
                fmt15(c["sdi"]),
                var_s))

    def _build_grafico_y_stats(self, parent, color):
        ciclos = self.hist.get(self.key,{}).get("ciclos",[])
        sdis   = [c["sdi"] for c in ciclos]

        # Stats resumen
        card = tk.Frame(parent, bg=C["surface"], padx=14, pady=12)
        card.pack(fill="x", pady=(0,10))
        tk.Label(card, text="ESTADÍSTICAS",
                 bg=C["surface"], fg=color,
                 font=("Georgia",7,"bold")).pack(anchor="w")
        tk.Frame(card, bg=C["border"], height=1).pack(
            fill="x", pady=(4,8))

        def stat(lbl, val, fg=None):
            r=tk.Frame(card,bg=C["surface"]); r.pack(fill="x",pady=2)
            tk.Label(r,text=lbl,bg=C["surface"],fg=C["text2"],
                     font=F["small"],width=14,anchor="w").pack(side="left")
            tk.Label(r,text=val,bg=C["surface"],
                     fg=fg or C["text"],
                     font=("Courier New",9,"bold")).pack(side="left")

        n = len(ciclos)
        stat("Ciclos:", f"{n} / {N_CICLOS}",
             fg=C["green"] if n>=N_CICLOS else C["yellow"])
        if sdis:
            stat("s(ΔI) mín.:", fmt15(min(sdis)), fg=C["green"])
            stat("s(ΔI) máx.:", fmt15(max(sdis)))
            stat("s(ΔI) media:", fmt15(sum(sdis)/len(sdis)))
        if len(sdis)>=2:
            var = varianza(sdis)
            if var is not None:
                stat("Varianza:", fmt15(var), fg=C["accent"])

        # Gráfico
        tk.Label(parent, text="TENDENCIA s(ΔI)",
                 bg=C["bg"], fg=color,
                 font=("Georgia",7,"bold")).pack(anchor="w",
                                                 pady=(8,0))
        tk.Frame(parent, bg=C["border"], height=1).pack(
            fill="x", pady=(4,6))
        fig = Figure(figsize=(3.4,3), facecolor=C["surface"])
        ax  = fig.add_subplot(111)
        ax.set_facecolor(C["surface2"])
        ax.tick_params(colors=C["text3"], labelsize=7)
        for sp in ["bottom","left"]:
            ax.spines[sp].set_color(C["border2"])
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.set_xlabel("Ciclo", color=C["text3"], fontsize=7)
        ax.set_ylabel("s(ΔI) g", color=C["text3"], fontsize=7)
        ax.grid(color=C["border"], linestyle="--", linewidth=0.4)

        if ciclos:
            xs=[c["n"] for c in ciclos]; ys=sdis
            ax.plot(xs,ys,color=color,linewidth=1.8,
                marker="o",markersize=5,
                markerfacecolor=C["green"],
                markeredgecolor=C["green"])
            if len(ys)>=2:
                media=sum(ys)/len(ys)
                v=varianza(ys); sigma=math.sqrt(v) if v else 0
                ax.axhline(media,color=C["yellow"],
                    linewidth=1,linestyle="--",label="media")
                ax.axhline(media+sigma,color=C["red"],
                    linewidth=0.7,linestyle=":",alpha=0.8,label="+σ")
                ax.axhline(media-sigma,color=C["red"],
                    linewidth=0.7,linestyle=":",alpha=0.8)
            if xs:
                ax.annotate(f"{ys[-1]:.5f}",(xs[-1],ys[-1]),
                    textcoords="offset points",
                    xytext=(5,5),fontsize=6,color=color)
            ax.set_xticks(xs)
            if len(ys)>=2:
                ax.legend(fontsize=6,labelcolor=C["text2"],
                    facecolor=C["surface2"],
                    edgecolor=C["border"],loc="upper right")
        else:
            ax.text(0.5,0.5,"Sin datos",transform=ax.transAxes,
                ha="center",va="center",color=C["text3"],fontsize=9)

        ax.set_title(f"s(ΔI)  {self.pesa[1]}",
            color=C["text2"],fontsize=8,pad=4)
        fig.tight_layout(pad=1.2)
        canvas = FigureCanvasTkAgg(fig, master=parent)
        canvas.get_tk_widget().pack(fill="both", expand=True)
        canvas.draw()

        # Botones exportar desde esta ventana
        btn_row = tk.Frame(parent, bg=C["bg"])
        btn_row.pack(fill="x", pady=(10,0))
        tk.Button(btn_row,
            text="  Excel  ",
            bg=color, fg="white",
            font=F["h3"], relief="flat",
            pady=6,
            command=lambda: exportar_excel(
                self.bkey, self.pesa, self.hist, self)
            ).pack(side="left", fill="x", expand=True, padx=(0,4))
        tk.Button(btn_row,
            text="  PDF  ",
            bg=self.bcfg["color_dark"], fg="white",
            font=F["h3"], relief="flat",
            pady=6,
            command=lambda: exportar_pdf(
                self.bkey, self.pesa, self.hist, self)
            ).pack(side="left", fill="x", expand=True)

    def _borrar(self):
        if messagebox.askyesno("Confirmar",
                f"¿Eliminar historial de {self.pesa[1]}?",
                parent=self):
            self.hist[self.key]={"ciclos":[]}
            guardar(self.hist)
            if self.on_borrar: self.on_borrar()
            self.destroy()


# ══════════════════════════════════════════════════════════════
#  TARJETA DE PESA — widget del dashboard
# ══════════════════════════════════════════════════════════════
class TarjetaPesa(tk.Frame):
    def __init__(self, parent, balanza_key, pesa, hist,
                 cx_getter, on_ciclo_guardado, **kw):
        super().__init__(parent, bg=C["surface"],
                         highlightthickness=1,
                         highlightbackground=C["border"],
                         **kw)
        self.bkey    = balanza_key
        self.bcfg    = BALANZAS_CFG[balanza_key]
        self.pesa    = pesa
        self.hist    = hist
        self.cx_getter = cx_getter
        self.on_ciclo_guardado = on_ciclo_guardado
        self._build()
        self.refresh()

    @property
    def key(self): return f"{self.bcfg['nombre']}_{self.pesa[1]}"

    def _build(self):
        color = self.bcfg["color_hex"]

        # ── Cabecera tarjeta ─────────────────────────────
        top = tk.Frame(self, bg=color, padx=12, pady=8)
        top.pack(fill="x")
        tk.Label(top, text=self.pesa[1],
                 bg=color, fg="white",
                 font=("Georgia",12,"bold")).pack(side="left")
        # Icono de alerta (se actualiza en refresh)
        self.lbl_alerta_icono = tk.Label(top, text="",
            bg=color, fg="white",
            font=("Georgia",11,"bold"))
        self.lbl_alerta_icono.pack(side="right")

        # ── Cuerpo ───────────────────────────────────────
        body = tk.Frame(self, bg=C["surface"], padx=12, pady=8)
        body.pack(fill="both", expand=True)

        # Ciclos ABA de la sesión activa
        prog_row = tk.Frame(body, bg=C["surface"])
        prog_row.pack(fill="x", pady=(0,2))
        tk.Label(prog_row, text="Ciclos ABA:",
                 bg=C["surface"], fg=C["text2"],
                 font=F["small"]).pack(side="left")
        self.lbl_ciclos = tk.Label(prog_row,
            text="0 / 10",
            bg=C["surface"], fg=C["text"],
            font=("Courier New",10,"bold"))
        self.lbl_ciclos.pack(side="left", padx=6)

        # Sesiones del año
        r0 = tk.Frame(body, bg=C["surface"]); r0.pack(fill="x", pady=(0,4))
        tk.Label(r0, text="Caract. año:",
                 bg=C["surface"], fg=C["text2"],
                 font=F["small"], width=12, anchor="w").pack(side="left")
        self.lbl_sesiones = tk.Label(r0, text="0 / 2",
            bg=C["surface"], fg=C["text3"],
            font=("Courier New",8,"bold"))
        self.lbl_sesiones.pack(side="left")

        # Barra de progreso ciclos
        self.canvas_prog = tk.Canvas(body, height=6,
            bg=C["surface3"], highlightthickness=0)
        self.canvas_prog.pack(fill="x", pady=(0,6))

        # s(ΔI) última
        r1=tk.Frame(body,bg=C["surface"]); r1.pack(fill="x",pady=1)
        tk.Label(r1,text="Ultimo s(DI):",bg=C["surface"],
                 fg=C["text2"],font=F["small"],
                 width=12,anchor="w").pack(side="left")
        self.lbl_sdi=tk.Label(r1,text="—",
            bg=C["surface"],fg=C["accent"],
            font=("Courier New",8,"bold"))
        self.lbl_sdi.pack(side="left")

        # Varianza entre ciclos
        r2=tk.Frame(body,bg=C["surface"]); r2.pack(fill="x",pady=1)
        tk.Label(r2,text="Varianza:",bg=C["surface"],
                 fg=C["text2"],font=F["small"],
                 width=12,anchor="w").pack(side="left")
        self.lbl_var=tk.Label(r2,text="—",
            bg=C["surface"],fg=C["purple"],
            font=("Courier New",8,"bold"))
        self.lbl_var.pack(side="left")

        # Estado / Alerta
        self.lbl_estado=tk.Label(body,text="Sin caracterizaciones",
            bg=C["surface"],fg=C["text3"],
            font=("Georgia",7,"italic"))
        self.lbl_estado.pack(anchor="w",pady=(4,0))

        sep(body, color=C["border2"], h=1, pady=6)

        # Botones
        btns=tk.Frame(body,bg=C["surface"]); btns.pack(fill="x")
        tk.Button(btns,
            text="▶  Nuevo ciclo ABA",
            bg=color, fg="white",
            font=("Georgia",8,"bold"),
            relief="flat",padx=8,pady=5,
            command=self._nuevo_ciclo).pack(side="left",fill="x",
                                            expand=True,padx=(0,4))
        tk.Button(btns,
            text="Historial",
            bg=C["surface3"],fg=C["text2"],
            font=F["small"],relief="flat",padx=6,pady=5,
            command=self._ver_historial).pack(side="left")

    def refresh(self):
        color = self.bcfg["color_hex"]
        anio  = datetime.now().year
        hist_key = self.hist.get(self.key, {})
        sesiones = hist_key.get("sesiones", [])

        # Sesión activa
        ses = sesion_actual(self.hist, self.key)
        ciclos_ses = ses["ciclos"] if ses else []
        n = len(ciclos_ses)

        # Sesiones del año
        sas   = sesiones_anio(self.hist, self.key, anio)
        n_sas = len(sas)
        color_sas = C["red"] if n_sas >= MAX_SESIONES_ANIO else C["text3"]
        self.lbl_sesiones.config(
            text=f"{n_sas} / {MAX_SESIONES_ANIO}",
            fg=color_sas)

        # Contador ciclos ABA de sesión activa
        self.lbl_ciclos.config(text=f"{n} / {N_CICLOS}",
            fg=C["green"] if n>=N_CICLOS else C["text"])

        # Barra progreso
        self.canvas_prog.update_idletasks()
        w = self.canvas_prog.winfo_width() or 200
        self.canvas_prog.delete("all")
        self.canvas_prog.create_rectangle(0,0,w,6,
            fill=C["surface3"],outline="")
        pct = min(n/N_CICLOS,1.0)
        if pct>0:
            fc = C["green"] if pct>=1 else color
            self.canvas_prog.create_rectangle(0,0,int(w*pct),6,
                fill=fc,outline="")

        # Última alerta activa de TODAS las sesiones
        ultima_alerta_nivel = ALERTA_NINGUNA
        for s in sesiones:
            for a in s.get("alertas",[]):
                nv = nivel_max(a.get("items",[]))
                if nv > ultima_alerta_nivel:
                    ultima_alerta_nivel = nv

        # Icono alerta en header
        icono = ICONO_ALERTA.get(ultima_alerta_nivel,"")
        ca    = COLOR_ALERTA.get(ultima_alerta_nivel, color)
        self.lbl_alerta_icono.config(
            text=icono if ultima_alerta_nivel>ALERTA_NINGUNA else "",
            fg=ca)

        # s(ΔI) y varianza
        if ciclos_ses:
            sdis=[c["sdi"] for c in ciclos_ses]
            sdi_color = COLOR_ALERTA.get(ultima_alerta_nivel, C["accent"])
            self.lbl_sdi.config(text=fmt15(sdis[-1])+" g",
                                fg=sdi_color)
            if len(sdis)>=2:
                var=varianza(sdis)
                if var is not None:
                    self.lbl_var.config(text=fmt15(var)+" g\u00b2")
        else:
            # Buscar último s(ΔI) de sesiones anteriores
            all_c = []
            for s in sesiones: all_c.extend(s.get("ciclos",[]))
            if all_c:
                self.lbl_sdi.config(
                    text=fmt15(all_c[-1]["sdi"])+" g (ant.)",
                    fg=C["text3"])
            else:
                self.lbl_sdi.config(text="—", fg=C["accent"])
                self.lbl_var.config(text="—", fg=C["purple"])

        # Estado
        if ultima_alerta_nivel == ALERTA_CRITICA:
            self.lbl_estado.config(
                text="🚨 ALERTA CRITICA — Ver historial",
                fg=C["red"])
        elif ultima_alerta_nivel == ALERTA_MODERADA:
            self.lbl_estado.config(
                text="⚠  Atencion — Revisar historial",
                fg=C["yellow"])
        elif n >= N_CICLOS:
            self.lbl_estado.config(
                text=f"Sesion {ses['id_sesion'] if ses else ''} completa",
                fg=C["green"])
        elif sesiones:
            self.lbl_estado.config(
                text=f"Ciclos ABA: {n}/{N_CICLOS}",
                fg=C["accent"])
        else:
            self.lbl_estado.config(
                text="Sin caracterizaciones registradas",
                fg=C["text3"])

    def _nuevo_ciclo(self):
        if self.bcfg["tipo"]!="manual":
            cx=self.cx_getter(self.bkey)
            if not cx or not getattr(cx,"activo",False):
                messagebox.showwarning("Sin conexión",
                    f"Conecta la {self.bcfg['nombre']} primero.")
                return
        win=VentanaCarac(
            self.winfo_toplevel(),
            self.bkey, self.pesa, self.hist,
            self.cx_getter,
            on_guardado=self._guardado)
        # Registrar la ventana para recibir datos del stream
        self._ventana_activa=win

    def _guardado(self):
        self.refresh()
        if self.on_ciclo_guardado:
            self.on_ciclo_guardado()

    def _ver_historial(self):
        VentanaDetalle(
            self.winfo_toplevel(),
            self.bkey, self.pesa, self.hist,
            on_borrar=self.refresh)

    def entregar_dato(self, valor):
        """Recibe un valor del stream de la balanza."""
        if hasattr(self,"_ventana_activa") and \
                self._ventana_activa.winfo_exists():
            self._ventana_activa.recibir(valor)


# ══════════════════════════════════════════════════════════════
#  PANEL BALANZA — sección del dashboard para una balanza
# ══════════════════════════════════════════════════════════════
class SeccionBalanza(tk.Frame):
    def __init__(self, parent, balanza_key, hist,
                 on_ciclo_guardado, **kw):
        super().__init__(parent, bg=C["bg"], **kw)
        self.bkey    = balanza_key
        self.bcfg    = BALANZAS_CFG[balanza_key]
        self.hist    = hist
        self.on_ciclo_guardado = on_ciclo_guardado
        self.cx      = None
        self.tarjetas= {}
        self._build()

    def get_cx(self): return self.cx

    def _build(self):
        color = self.bcfg["color_hex"]
        tipo  = self.bcfg["tipo"]

        # ── Header sección ───────────────────────────────
        hdr = tk.Frame(self, bg=C["surface"], padx=16, pady=10)
        hdr.pack(fill="x", pady=(0,2))

        # Franja color izquierda
        tk.Frame(hdr, bg=color, width=4).pack(side="left", fill="y",
                                               padx=(0,12))
        info = tk.Frame(hdr, bg=C["surface"])
        info.pack(side="left", fill="both", expand=True)
        tk.Label(info,
                 text=f"{self.bcfg['nombre']}",
                 bg=C["surface"], fg=color,
                 font=("Georgia",12,"bold")).pack(anchor="w")
        tk.Label(info,
                 text=f"{self.bcfg['id_bal']}  |  Resolucion: {self.bcfg['resolucion']}",
                 bg=C["surface"], fg=C["text3"],
                 font=F["small"]).pack(anchor="w")

        # Panel conexión (lado derecho del header)
        cx_frame = tk.Frame(hdr, bg=C["surface"])
        cx_frame.pack(side="right")
        self._build_conexion(cx_frame, color, tipo)

        # ── Grid de tarjetas ─────────────────────────────
        grid = tk.Frame(self, bg=C["bg"])
        grid.pack(fill="x", padx=0, pady=6)

        pesas = self.bcfg["pesas"]
        cols  = 4 if len(pesas)>3 else len(pesas)
        for i, pesa in enumerate(pesas):
            c_idx = i % cols
            r_idx = i // cols
            grid.columnconfigure(c_idx, weight=1, uniform="card")
            card = TarjetaPesa(
                grid, self.bkey, pesa, self.hist,
                cx_getter=lambda k=self.bkey: self.get_cx(),
                on_ciclo_guardado=self.on_ciclo_guardado)
            card.grid(row=r_idx, column=c_idx,
                      sticky="nsew", padx=4, pady=4)
            self.tarjetas[pesa[1]] = card

    def _build_conexion(self, parent, color, tipo):
        if tipo=="serial":
            tk.Label(parent,text="Puerto:",bg=C["surface"],
                     fg=C["text2"],font=F["small"]).pack(side="left")
            self.combo_port=ttk.Combobox(parent,width=6,
                state="readonly")
            puertos=[x.device for x in
                     serial.tools.list_ports.comports()]
            self.combo_port["values"]=puertos
            dflt=self.bcfg.get("puerto","COM6")
            self.combo_port.set(
                dflt if dflt in puertos
                else (puertos[0] if puertos else ""))
            self.combo_port.pack(side="left",padx=4)
            tk.Label(parent,text="Baud:",bg=C["surface"],
                     fg=C["text2"],font=F["small"]).pack(side="left")
            self.combo_baud=ttk.Combobox(parent,width=5,
                state="readonly",
                values=["9600","19200","4800","2400"])
            self.combo_baud.set(str(self.bcfg.get("baud",9600)))
            self.combo_baud.pack(side="left",padx=4)
            tk.Button(parent,text="↺",bg=C["surface3"],
                fg=C["text2"],font=("Georgia",9),relief="flat",
                command=self._refresh_ports).pack(side="left",padx=2)
            self.btn_cx=tk.Button(parent,text="Conectar",
                bg=color,fg="white",font=F["h3"],
                relief="flat",padx=10,pady=4,
                command=self._toggle_cx)
            self.btn_cx.pack(side="left",padx=6)
            self.lbl_cx=tk.Label(parent,text="⚫ Desconectado",
                bg=C["surface"],fg=C["red"],font=F["small"])
            self.lbl_cx.pack(side="left")

        elif tipo=="wifi":
            tk.Label(parent,text="IP:",bg=C["surface"],
                     fg=C["text2"],font=F["small"]).pack(side="left")
            self.e_ip=tk.Entry(parent,width=13,
                font=("Courier New",8),bg=C["surface3"],
                fg=C["text"],insertbackground=color,
                relief="flat",bd=0,highlightthickness=1,
                highlightbackground=C["border"],
                highlightcolor=color)
            self.e_ip.insert(0,self.bcfg.get("ip",RADWAG_IP))
            self.e_ip.pack(side="left",padx=4,ipady=3)
            tk.Label(parent,text="Puerto:",bg=C["surface"],
                     fg=C["text2"],font=F["small"]).pack(side="left")
            self.e_port=tk.Entry(parent,width=5,
                font=("Courier New",8),bg=C["surface3"],
                fg=C["text"],insertbackground=color,
                relief="flat",bd=0,highlightthickness=1,
                highlightbackground=C["border"],
                highlightcolor=color)
            self.e_port.insert(0,str(self.bcfg.get("port",RADWAG_PORT)))
            self.e_port.pack(side="left",padx=4,ipady=3)
            self.btn_cx=tk.Button(parent,text="Conectar",
                bg=color,fg="white",font=F["h3"],
                relief="flat",padx=10,pady=4,
                command=self._toggle_cx)
            self.btn_cx.pack(side="left",padx=6)
            self.lbl_cx=tk.Label(parent,text="⚫ Desconectado",
                bg=C["surface"],fg=C["red"],font=F["small"])
            self.lbl_cx.pack(side="left")

        elif tipo=="manual":
            self.lbl_cx=tk.Label(parent,
                text="🟠 Ingreso manual activo",
                bg=C["surface"],fg=self.bcfg["color_hex"],
                font=F["small"])
            self.lbl_cx.pack(side="left")

    def _refresh_ports(self):
        p=[x.device for x in serial.tools.list_ports.comports()]
        self.combo_port["values"]=p

    def _toggle_cx(self):
        color=self.bcfg["color_hex"]; tipo=self.bcfg["tipo"]
        if self.cx and getattr(self.cx,"activo",False):
            self.cx.desconectar(); self.cx=None
            self.btn_cx.config(text="Conectar",bg=color)
            self.lbl_cx.config(text="⚫ Desconectado",fg=C["red"])
        else:
            if tipo=="serial":
                puerto=self.combo_port.get()
                baud=int(self.combo_baud.get())
                self.cx=ConexionSerial(cb=self._on_dato)
                ok=self.cx.conectar(puerto,baud)
                if ok:
                    self.btn_cx.config(text="Desconectar",bg=C["red"])
                    self.lbl_cx.config(
                        text=f"🟢 {puerto} @ {baud}",fg=C["green"])
                else:
                    messagebox.showerror("Error",
                        f"No se pudo abrir {puerto}.")
                    self.cx=None
            elif tipo=="wifi":
                ip=self.e_ip.get().strip()
                port=int(self.e_port.get().strip())
                self.cx=ConexionRadwag(cb=self._on_dato)
                self.cx.conectar(ip,port)
                self.btn_cx.config(text="Desconectar",bg=C["red"])
                self.lbl_cx.config(
                    text=f"🟢 WiFi {ip}:{port}",fg=C["green"])

    def _on_dato(self,valor,raw):
        self.after(0,self._procesar,valor,raw)

    def _procesar(self,valor,raw):
        if raw=="__CONNECTED__":
            self.lbl_cx.config(text="🟢 WiFi conectado",fg=C["green"]); return
        if raw=="__DISCONNECTED__":
            self.lbl_cx.config(text="🔄 Reconectando...",fg=C["yellow"]); return
        if valor is None: return
        # Entregar al tarjeta que tenga ventana activa
        for card in self.tarjetas.values():
            card.entregar_dato(valor)

    def refresh_all(self):
        for card in self.tarjetas.values():
            card.refresh()


# ══════════════════════════════════════════════════════════════
#  EXPORTACIÓN EXCEL
# ══════════════════════════════════════════════════════════════
def exportar_excel(balanza_key, pesa, hist, parent):
    try:
        import openpyxl
        from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
    except ImportError:
        messagebox.showerror("Error","pip install openpyxl",parent=parent)
        return

    bcfg = BALANZAS_CFG[balanza_key]
    key  = f"{bcfg['nombre']}_{pesa[1]}"
    ciclos = hist.get(key,{}).get("ciclos",[])
    if not ciclos:
        messagebox.showinfo("Sin datos","No hay ciclos.",parent=parent); return

    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta = filedialog.asksaveasfilename(
        parent=parent,
        defaultextension=".xlsx",
        filetypes=[("Excel","*.xlsx")],
        initialfile=f"carac_{key.replace(' ','_')}_{ts}.xlsx")
    if not ruta: return

    def fill(h): return PatternFill("solid",fgColor=h)
    def font(bold=False,color="000000",sz=10):
        return Font(bold=bold,color=color,size=sz)
    def aln(h="center",wrap=False):
        return Alignment(horizontal=h,vertical="center",wrap_text=wrap)
    def brd():
        s=Side(style="thin",color="CCCCCC")
        return Border(left=s,right=s,top=s,bottom=s)
    def brd_m():
        s=Side(style="medium",color=bcfg["color_hex"].replace("#",""))
        return Border(left=s,right=s,top=s,bottom=s)

    wb=openpyxl.Workbook(); sdis=[c["sdi"] for c in ciclos]
    ords={1:"1er",2:"2do",3:"3er",4:"4to",5:"5to",
          6:"6to",7:"7mo",8:"8vo",9:"9no",10:"10mo"}

    for idx_c,ciclo in enumerate(ciclos):
        n_c=ciclo["n"]
        ws=wb.active if idx_c==0 else wb.create_sheet(f"Ciclo {n_c}")
        ws.title=f"Ciclo {n_c}"
        for col,w in zip("ABCDE",[6,16,16,16,22]):
            ws.column_dimensions[col].width=w
        ws.column_dimensions["G"].width=48

        # Fila 1
        ws.merge_cells("A1:E1"); c=ws["A1"]
        c.value=f"METROMECANICA  |  {bcfg['id_bal']}  —  {bcfg['nombre']}  |  NMP 004:2007  |  ISO/IEC 17025"
        c.font=font(True,"FFFFFF",10); c.fill=fill("17375E")
        c.alignment=aln(); c.border=brd(); ws.row_dimensions[1].height=16

        # Fila 2 — patrón
        ws.merge_cells("A2:E2"); c=ws["A2"]
        c.value=(f"Pesa: {pesa[1]}  |  "
                 f"ID: {ciclo.get('pat_id','—')}  |  "
                 f"Cert.: {ciclo.get('pat_cert','—')}  |  "
                 f"Fecha: {ciclo.get('pat_fecha','—')}")
        c.font=font(color="404040",sz=9); c.fill=fill("F2F2F2")
        c.alignment=aln("left"); c.border=brd(); ws.row_dimensions[2].height=14

        # Fila 3 — condiciones ambientales
        ci=ciclo.get("cond_inicio",{}); cf=ciclo.get("cond_fin",{})
        ws.merge_cells("A3:E3"); c=ws["A3"]
        def _cv(d,k): return d.get(k,"—") or "—"
        c.value=(
            f"INICIO:  T={_cv(ci,'temp')} C  |  "
            f"HR={_cv(ci,'hum')} %  |  "
            f"P={_cv(ci,'pres')} hPa          "
            f"FIN:  T={_cv(cf,'temp')} C  |  "
            f"HR={_cv(cf,'hum')} %  |  "
            f"P={_cv(cf,'pres')} hPa"
        )
        c.font=font(color="1a4040",sz=8); c.fill=fill("E6F2F2")
        c.alignment=aln("left"); c.border=brd(); ws.row_dimensions[3].height=13

        # Fila 4 — título: "Nra Caracterización" no "Ciclo"
        ords_carac = {1:"1ra",2:"2da",3:"3ra",4:"4ta",5:"5ta",
                      6:"6ta",7:"7ma",8:"8va",9:"9na",10:"10ma"}
        ws.merge_cells("A4:E4"); c=ws["A4"]
        c.value=(f"Caracterizacion de la Balanza  —  "
                 f"Pesa {pesa[1]}  —  "
                 f"{ords_carac.get(n_c,str(n_c))} Caracterizacion")
        c.font=font(True,sz=11); c.fill=fill("BDD7EE")
        c.alignment=aln(); c.border=brd(); ws.row_dimensions[4].height=18

        # Fila 5 — Cabeceras
        for txt,col in [("N","A"),("PATRON (Ir1)\ng","B"),
                        ("CALIBRAR (It)\ng","C"),
                        ("PATRON (Ir2)\ng","D"),("DI\ng","E")]:
            cel=ws[f"{col}5"]; cel.value=txt
            cel.font=font(True,sz=10); cel.fill=fill("BDD7EE")
            cel.alignment=aln("center",wrap=True); cel.border=brd()
        ws.row_dimensions[5].height=28

        # Lecturas desde fila 6
        # dd = decimales de la balanza (1 para WANT, 2 BIOBASE, 4 RADWAG)
        # DI se guarda con precision COMPLETA (14 decimales) sin redondear
        dd=pesa[3]; lects=ciclo.get("lecturas",[])
        for j,l in enumerate(lects,1):
            fila=5+j
            ir1=l["ir1"]; it=l["it"]; ir2=l["ir2"]
            di=delta_i(ir1,it,ir2)
            for col,v in zip(["A","B","C","D","E"],[j,ir1,it,ir2,di]):
                cel=ws[f"{col}{fila}"]
                if col=="A":
                    cel.value=v
                elif col=="E":
                    # ΔI: valor exacto, formato con 14 decimales
                    cel.value=v
                    cel.number_format="0."+"0"*14
                    cel.fill=fill("E2EFDA")
                else:
                    # Ir1, It, Ir2: redondear a decimales de la balanza
                    cel.value=round(v,dd)
                    cel.number_format="0."+"0"*dd
                cel.font=font(sz=10); cel.alignment=aln("center"); cel.border=brd()
            ws.row_dimensions[fila].height=15

        fila_s=5+len(lects)+2
        ws.merge_cells(f"A{fila_s}:D{fila_s}"); c=ws[f"A{fila_s}"]
        c.value="s(ΔI) :  Desviacion estandar de las diferencias de Lectura de la pesa a calibrar y la pesa de referencia"
        c.font=font(sz=9,color="404040"); c.alignment=aln("right")
        ws.row_dimensions[fila_s].height=26
        c_s=ws[f"E{fila_s}"]
        # s(ΔI): valor exacto con 13 decimales (coincide con tu referencia)
        c_s.value=ciclo["sdi"]; c_s.font=font(True,sz=11)
        c_s.fill=fill("00B0F0"); c_s.alignment=aln("center")
        c_s.border=brd_m(); c_s.number_format="0.0000000000000"  # 13 decimales

        fila_v=fila_s+1
        ws.merge_cells(f"A{fila_v}:C{fila_v}"); c=ws[f"A{fila_v}"]
        c.value="Varianza de DI"; c.font=font(True,sz=10)
        c.fill=fill("D9EAD3"); c.alignment=aln("right"); c.border=brd()
        vac=varianza(sdis[:n_c]) if n_c>=2 else None
        c_v=ws[f"E{fila_v}"]
        if vac is not None:
            c_v.value=vac; c_v.font=font(True,sz=11,color="FFFFFF")
            c_v.fill=fill("0D9488"); c_v.number_format="0.0000000000000"
        else:
            c_v.value="--- (min. 2 caracterizaciones)"
            c_v.font=font(sz=9,color="808080"); c_v.fill=fill("EEEEEE")
        c_v.alignment=aln("center"); c_v.border=brd_m()
        ws[f"D{fila_v}"].value="Para 2 o mas caracterizaciones"
        ws[f"D{fila_v}"].font=font(sz=8,color="606060")
        ws[f"D{fila_v}"].alignment=aln("left")
        ws.row_dimensions[fila_v].height=20

    # Hoja resumen
    if len(ciclos)>1:
        ws_r=wb.create_sheet("Resumen")
        for col,w in zip("ABCDE",[8,20,22,22,22]):
            ws_r.column_dimensions[col].width=w
        ws_r.merge_cells("A1:E1"); c=ws_r["A1"]
        c.value=f"RESUMEN  {bcfg['id_bal']}  —  {bcfg['nombre']}  |  Pesa {pesa[1]}"
        c.font=font(True,"FFFFFF",11); c.fill=fill("17375E")
        c.alignment=aln(); ws_r.row_dimensions[1].height=18
        for txt,col in [("Ciclo","A"),("Fecha","B"),("ID Patrón","C"),
                        ("s(ΔI) g","D"),("Varianza g²","E")]:
            cel=ws_r[f"{col}2"]; cel.value=txt
            cel.font=font(True); cel.fill=fill("BDD7EE")
            cel.border=brd(); cel.alignment=aln()
        ws_r.row_dimensions[2].height=16
        for i,c in enumerate(ciclos,1):
            fila=2+i
            vac=varianza(sdis[:i]) if i>=2 else None
            for col,v in zip(["A","B","C","D","E"],
                    [c["n"],c["fecha"],c.get("pat_id","—"),
                     c["sdi"],vac if vac is not None else "—"]):
                cel=ws_r[f"{col}{fila}"]
                cel.value=v; cel.font=font(sz=10)
                cel.alignment=aln("center"); cel.border=brd()
                if col=="D" and isinstance(v,float): cel.number_format="0.0000000000000"
                if col=="E" and isinstance(v,float): cel.number_format="0.0000000000000"
            ws_r.row_dimensions[fila].height=14

    wb.save(ruta)
    messagebox.showinfo("Excel exportado",f"Guardado:\n{ruta}",parent=parent)
    if messagebox.askyesno("Abrir","¿Abrir ahora?",parent=parent):
        abrir(ruta)


# ══════════════════════════════════════════════════════════════
#  EXPORTACIÓN PDF
# ══════════════════════════════════════════════════════════════
def exportar_pdf(balanza_key, pesa, hist, parent):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (SimpleDocTemplate, Table,
            TableStyle, Paragraph, Spacer, HRFlowable)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        messagebox.showerror("Error","pip install reportlab",parent=parent)
        return

    bcfg=BALANZAS_CFG[balanza_key]
    key=f"{bcfg['nombre']}_{pesa[1]}"
    ciclos=hist.get(key,{}).get("ciclos",[])
    if not ciclos:
        messagebox.showinfo("Sin datos","No hay ciclos.",parent=parent); return

    ts=datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta=filedialog.asksaveasfilename(
        parent=parent,
        defaultextension=".pdf",
        filetypes=[("PDF","*.pdf")],
        initialfile=f"carac_{key.replace(' ','_')}_{ts}.pdf")
    if not ruta: return

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate,Table,
        TableStyle,Paragraph,Spacer,HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet,ParagraphStyle

    # Color de la balanza como objeto reportlab
    hex_c=bcfg["color_hex"].lstrip("#")
    r,g,b=(int(hex_c[i:i+2],16)/255 for i in (0,2,4))
    accent_col=colors.Color(r,g,b)
    dark_col=colors.Color(0.07,0.12,0.22)

    doc=SimpleDocTemplate(ruta,pagesize=A4,
        leftMargin=1.8*cm,rightMargin=1.8*cm,
        topMargin=1.5*cm,bottomMargin=1.5*cm)

    styles=getSampleStyleSheet()
    def st(name,**kw):
        return ParagraphStyle(name,parent=styles["Normal"],**kw)

    s_titulo=st("t",fontSize=18,fontName="Helvetica-Bold",
        textColor=colors.white,leading=22)
    s_sub=st("s",fontSize=9,fontName="Helvetica",
        textColor=colors.HexColor("#94a3b8"),leading=12)
    s_head=st("h",fontSize=10,fontName="Helvetica-Bold",
        textColor=accent_col,leading=14,spaceAfter=4)
    s_body=st("b",fontSize=9,fontName="Helvetica",
        textColor=colors.HexColor("#e2e8f0"),leading=12)
    s_mono=st("m",fontSize=8,fontName="Courier",
        textColor=colors.HexColor("#e2e8f0"),leading=11)

    story=[]
    sdis=[c["sdi"] for c in ciclos]
    W=A4[0]-3.6*cm

    # ── Para cada ciclo ─────────────────────────────────
    for idx_c,ciclo in enumerate(ciclos):
        n_c=ciclo["n"]
        if idx_c>0:
            from reportlab.platypus import PageBreak
            story.append(PageBreak())

        # Cabecera
        hdr_data=[
            [Paragraph("METROMECANICA",s_titulo),
             Paragraph(
                 f"{bcfg['id_bal']}  —  {bcfg['nombre']}<br/>"
                 f"NMP 004:2007  |  ISO/IEC 17025",
                 st("x",fontSize=8,fontName="Helvetica",
                     textColor=colors.HexColor("#94a3b8"),
                     leading=12,alignment=2))],
        ]
        hdr_tbl=Table(hdr_data,colWidths=[W*0.55,W*0.45])
        hdr_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),dark_col),
            ("TOPPADDING",(0,0),(-1,-1),14),
            ("BOTTOMPADDING",(0,0),(-1,-1),14),
            ("LEFTPADDING",(0,0),(0,-1),16),
            ("RIGHTPADDING",(-1,0),(-1,-1),16),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ]))
        story.append(hdr_tbl)
        story.append(Spacer(1,8))

        # Info caracterizacion
        ords_carac={1:"1ra",2:"2da",3:"3ra",4:"4ta",5:"5ta",
                    6:"6ta",7:"7ma",8:"8va",9:"9na",10:"10ma"}
        story.append(Paragraph(
            f"Caracterizacion de Balanza  —  Pesa <b>{pesa[1]}</b>  —  "
            f"{ords_carac.get(n_c,str(n_c))} Caracterizacion",
            st("ci",fontSize=11,fontName="Helvetica-Bold",
               textColor=accent_col,leading=14,spaceAfter=4)))
        story.append(Paragraph(
            f"{bcfg['id_bal']}  —  {bcfg['nombre']}  |  "
            f"Fecha: {ciclo.get('fecha','—')}  |  "
            f"ID Patron: {ciclo.get('pat_id','—')}  |  "
            f"Cert.: {ciclo.get('pat_cert','—')}",
            s_body))

        # Condiciones ambientales
        ci=ciclo.get("cond_inicio",{}); cf=ciclo.get("cond_fin",{})
        def cv(d,k): return d.get(k,"—") or "—"
        cond_data=[
            ["CONDICIONES INICIO",
             f"T: {cv(ci,'temp')} C   HR: {cv(ci,'hum')} %   P: {cv(ci,'pres')} hPa",
             "CONDICIONES FIN",
             f"T: {cv(cf,'temp')} C   HR: {cv(cf,'hum')} %   P: {cv(cf,'pres')} hPa"],
        ]
        cond_tbl=Table(cond_data,colWidths=[W*0.18,W*0.32,W*0.18,W*0.32])
        cond_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(0,-1),dark_col),
            ("BACKGROUND",(2,0),(2,-1),dark_col),
            ("TEXTCOLOR",(0,0),(0,-1),accent_col),
            ("TEXTCOLOR",(2,0),(2,-1),accent_col),
            ("FONTNAME",(0,0),(-1,-1),"Helvetica"),
            ("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),
            ("FONTNAME",(2,0),(2,-1),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),8),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("TOPPADDING",(0,0),(-1,-1),5),
            ("BOTTOMPADDING",(0,0),(-1,-1),5),
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#243044")),
            ("TEXTCOLOR",(1,0),(1,-1),colors.HexColor("#e2e8f0")),
            ("TEXTCOLOR",(3,0),(3,-1),colors.HexColor("#e2e8f0")),
        ]))
        story.append(Spacer(1,6))
        story.append(cond_tbl)
        story.append(Spacer(1,10))
        story.append(HRFlowable(width=W,color=accent_col,
            thickness=0.8,spaceAfter=10))

        # Tabla de lecturas
        dd=pesa[3]
        tbl_hdr=[["N°","PATRÓN Ir1 (g)","CALIBRAR It (g)",
                  "PATRÓN Ir2 (g)","ΔI (g)"]]
        lects=ciclo.get("lecturas",[])
        tbl_rows=[]
        for j,l in enumerate(lects,1):
            ir1=l["ir1"]; it=l["it"]; ir2=l["ir2"]
            di=delta_i(ir1,it,ir2)
            tbl_rows.append([
                str(j),
                fmt(ir1,dd), fmt(it,dd),
                fmt(ir2,dd), fmt(di,dd)])

        tbl_data=tbl_hdr+tbl_rows
        col_ws=[0.6*cm]+[((W-0.6*cm)/4)]*4
        t=Table(tbl_data,colWidths=col_ws,repeatRows=1)
        ts_=TableStyle([
            # Header
            ("BACKGROUND",(0,0),(-1,0),dark_col),
            ("TEXTCOLOR",(0,0),(-1,0),accent_col),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,0),8),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("TOPPADDING",(0,0),(-1,-1),5),
            ("BOTTOMPADDING",(0,0),(-1,-1),5),
            # Filas alternas
            ("BACKGROUND",(0,1),(-1,-1),colors.HexColor("#111827")),
            ("TEXTCOLOR",(0,1),(-1,-1),colors.HexColor("#e2e8f0")),
            ("FONTNAME",(0,1),(-1,-1),"Courier"),
            ("FONTSIZE",(0,1),(-1,-1),8),
            # ΔI columna resaltada
            ("BACKGROUND",(4,1),(4,-1),colors.HexColor("#1a2e1a")),
            ("TEXTCOLOR",(4,1),(4,-1),colors.HexColor("#10b981")),
            # Grid
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#243044")),
            ("LINEBELOW",(0,0),(-1,0),1,accent_col),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),
             [colors.HexColor("#111827"),colors.HexColor("#1a2235")]),
        ])
        t.setStyle(ts_)
        story.append(t)
        story.append(Spacer(1,14))

        # Resultados s(ΔI) y varianza
        res_data=[
            ["s(ΔI)  —  Desviación estándar de ΔI:",
             Paragraph(f"<b>{fmt15(ciclo['sdi'])} g</b>",
                       st("r",fontSize=10,fontName="Courier-Bold",
                           textColor=colors.HexColor("#38bdf8")))],
        ]
        vac=varianza(sdis[:n_c]) if n_c>=2 else None
        if vac is not None:
            res_data.append([
                "Varianza de ΔI entre ciclos:",
                Paragraph(f"<b>{fmt15(vac)} g\xb2</b>",
                          st("rv",fontSize=10,fontName="Courier-Bold",
                              textColor=colors.HexColor("#10b981")))])
        else:
            res_data.append([
                "Varianza de ΔI:",
                Paragraph("— (mínimo 2 ciclos)",
                          st("rv2",fontSize=9,fontName="Helvetica",
                              textColor=colors.HexColor("#4a5568")))])

        res_tbl=Table(res_data,colWidths=[W*0.55,W*0.45])
        res_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),dark_col),
            ("TEXTCOLOR",(0,0),(0,-1),colors.HexColor("#94a3b8")),
            ("FONTNAME",(0,0),(0,-1),"Helvetica"),
            ("FONTSIZE",(0,0),(0,-1),9),
            ("ALIGN",(0,0),(0,-1),"RIGHT"),
            ("ALIGN",(1,0),(1,-1),"LEFT"),
            ("TOPPADDING",(0,0),(-1,-1),8),
            ("BOTTOMPADDING",(0,0),(-1,-1),8),
            ("LEFTPADDING",(0,0),(0,-1),12),
            ("RIGHTPADDING",(0,0),(0,-1),12),
            ("LEFTPADDING",(1,0),(1,-1),12),
            ("LINEABOVE",(0,0),(-1,0),1,accent_col),
            ("LINEBELOW",(0,-1),(-1,-1),1,accent_col),
            ("LINEBEFORE",(0,0),(0,-1),3,accent_col),
        ]))
        story.append(res_tbl)

    # ── Hoja resumen final ──────────────────────────────
    from reportlab.platypus import PageBreak
    story.append(PageBreak())
    story.append(Paragraph("RESUMEN DE CARACTERIZACION",
        st("rs",fontSize=14,fontName="Helvetica-Bold",
           textColor=accent_col,spaceAfter=6)))
    story.append(Paragraph(
        f"{bcfg['id_bal']}  —  {bcfg['nombre']}  |  "
        f"Pesa {pesa[1]}  |  "
        f"{len(ciclos)} caracterizacion(es) registrada(s)",s_body))
    story.append(Spacer(1,10))
    story.append(HRFlowable(width=W,color=accent_col,
        thickness=0.8,spaceAfter=10))

    sum_hdr=[["N","Fecha","ID Patron","s(DI) g","Varianza g2"]]
    sum_rows=[]
    for i,c in enumerate(ciclos,1):
        vac=varianza(sdis[:i]) if i>=2 else None
        sum_rows.append([
            str(c["n"]), c["fecha"],
            c.get("pat_id","—"),
            fmt15(c["sdi"]),
            fmt15(vac) if vac else "—"])

    sum_data=sum_hdr+sum_rows
    col_ws2=[1*cm,3.5*cm,3*cm,5*cm,5*cm]
    sum_tbl=Table(sum_data,colWidths=col_ws2,repeatRows=1)
    sum_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),dark_col),
        ("TEXTCOLOR",(0,0),(-1,0),accent_col),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,0),8),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1),5),
        ("BOTTOMPADDING",(0,0),(-1,-1),5),
        ("BACKGROUND",(0,1),(-1,-1),colors.HexColor("#111827")),
        ("TEXTCOLOR",(0,1),(-1,-1),colors.HexColor("#e2e8f0")),
        ("FONTNAME",(0,1),(-1,-1),"Courier"),
        ("FONTSIZE",(0,1),(-1,-1),7),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),
         [colors.HexColor("#111827"),colors.HexColor("#1a2235")]),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#243044")),
        ("LINEBELOW",(0,0),(-1,0),1,accent_col),
    ]))
    story.append(sum_tbl)
    story.append(Spacer(1,16))

    # Firma final
    pie_data=[[
        Paragraph(
            f"METROMECANICA  |  NMP 004:2007  |  ISO/IEC 17025<br/>"
            f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            st("pie",fontSize=7,fontName="Helvetica",
               textColor=colors.HexColor("#4a5568"),leading=10))
    ]]
    pie_tbl=Table(pie_data,colWidths=[W])
    pie_tbl.setStyle(TableStyle([
        ("LINEABOVE",(0,0),(-1,0),0.5,colors.HexColor("#243044")),
        ("TOPPADDING",(0,0),(-1,-1),8),
    ]))
    story.append(pie_tbl)

    doc.build(story)
    messagebox.showinfo("PDF exportado",f"Guardado:\n{ruta}",parent=parent)
    if messagebox.askyesno("Abrir","¿Abrir ahora?",parent=parent):
        abrir(ruta)


# ══════════════════════════════════════════════════════════════
#  APLICACIÓN PRINCIPAL
# ══════════════════════════════════════════════════════════════
class App:
    def __init__(self, root):
        self.root=root
        self.root.title(
            "METROMECANICA — Sistema de Caracterización de Balanzas"
            " | NMP 004:2007 | ISO/IEC 17025")
        self.root.geometry("1440x900")
        self.root.configure(bg=C["bg"])
        self.root.minsize(1100,700)
        self.root.protocol("WM_DELETE_WINDOW",self._cerrar)

        self.hist=cargar()
        self.secciones={}
        self._build()
        self._tick()

    def _cerrar(self):
        for s in self.secciones.values():
            if s.cx and getattr(s.cx,"activo",False):
                s.cx.desconectar()
        self.root.destroy()

    def _build(self):
        # ── Header ─────────────────────────────────────
        hdr=tk.Frame(self.root,bg=C["surface"],padx=20,pady=10)
        hdr.pack(fill="x")
        tk.Frame(hdr,bg=C["accent"],width=3).pack(
            side="left",fill="y",padx=(0,12))
        left=tk.Frame(hdr,bg=C["surface"]); left.pack(side="left")
        tk.Label(left,text="METROMECANICA",
                 bg=C["surface"],fg=C["accent"],
                 font=("Georgia",16,"bold")).pack(anchor="w")
        tk.Label(left,
                 text="Sistema de Caracterización de Balanzas  "
                      "|  NMP 004:2007  |  Clase M2  |  ISO/IEC 17025",
                 bg=C["surface"],fg=C["text3"],
                 font=("Georgia",8,"italic")).pack(anchor="w")
        right=tk.Frame(hdr,bg=C["surface"]); right.pack(side="right")
        self.lbl_reloj=tk.Label(right,
            bg=C["surface"],fg=C["text2"],
            font=("Courier New",10))
        self.lbl_reloj.pack(anchor="e")
        # Botones exportar globales
        brow=tk.Frame(right,bg=C["surface"]); brow.pack(pady=(4,0))
        tk.Button(brow,text="Excel",
            bg=C["surface3"],fg=C["text2"],
            font=F["small"],relief="flat",padx=10,pady=4,
            command=self._exportar_xlsx).pack(side="left",padx=2)
        tk.Button(brow,text="PDF",
            bg=C["surface3"],fg=C["text2"],
            font=F["small"],relief="flat",padx=10,pady=4,
            command=self._exportar_pdf).pack(side="left",padx=2)

        tk.Frame(self.root,bg=C["border"],height=1).pack(fill="x")

        # ── Cuerpo scrollable ───────────────────────────
        canvas=tk.Canvas(self.root,bg=C["bg"],
                         highlightthickness=0)
        sb=ttk.Scrollbar(self.root,orient="vertical",
                         command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right",fill="y")
        canvas.pack(side="left",fill="both",expand=True)

        self.main=tk.Frame(canvas,bg=C["bg"])
        win_id=canvas.create_window((0,0),window=self.main,
                                     anchor="nw")

        def _resize(e):
            canvas.itemconfig(win_id,width=canvas.winfo_width())
        def _scroll(e):
            canvas.configure(scrollregion=canvas.bbox("all"))

        canvas.bind("<Configure>",_resize)
        self.main.bind("<Configure>",_scroll)
        canvas.bind_all("<MouseWheel>",
            lambda e: canvas.yview_scroll(-1*(e.delta//120),"units"))

        # ── Secciones por balanza ───────────────────────
        for bkey in ["RADWAG","BIOBASE","WANT"]:
            sec=SeccionBalanza(
                self.main, bkey, self.hist,
                on_ciclo_guardado=lambda: None)
            sec.pack(fill="x",padx=16,pady=(12,0))
            self.secciones[bkey]=sec
            tk.Frame(self.main,bg=C["border"],
                     height=1).pack(fill="x",padx=16,pady=(12,0))

    def _get_seleccion(self):
        """Dialogo para seleccionar balanza y pesa antes de exportar."""
        win=tk.Toplevel(self.root)
        win.title("Seleccionar para exportar")
        win.geometry("380x230")
        win.configure(bg=C["surface"])
        win.grab_set()
        win.resizable(False,False)
        # Centrar
        win.update_idletasks()
        x=self.root.winfo_x()+(self.root.winfo_width()-380)//2
        y=self.root.winfo_y()+(self.root.winfo_height()-230)//2
        win.geometry(f"+{x}+{y}")

        tk.Label(win,text="Selecciona que exportar",
                 bg=C["surface"],fg=C["text"],
                 font=F["h2"]).pack(pady=(16,4))
        tk.Frame(win,bg=C["border"],height=1).pack(
            fill="x",padx=20,pady=4)

        r1=tk.Frame(win,bg=C["surface"]); r1.pack(fill="x",padx=20,pady=4)
        tk.Label(r1,text="Balanza:",bg=C["surface"],
                 fg=C["text2"],font=F["body"],
                 width=10,anchor="w").pack(side="left")
        bal_var=tk.StringVar(value="RADWAG")
        cb_bal=ttk.Combobox(r1,textvariable=bal_var,
            values=["RADWAG","BIOBASE","WANT"],
            state="readonly",width=18)
        cb_bal.pack(side="left",padx=4)

        r2=tk.Frame(win,bg=C["surface"]); r2.pack(fill="x",padx=20,pady=4)
        tk.Label(r2,text="Pesa:",bg=C["surface"],
                 fg=C["text2"],font=F["body"],
                 width=10,anchor="w").pack(side="left")
        pesa_var=tk.StringVar()
        cb_pesa=ttk.Combobox(r2,textvariable=pesa_var,
            state="readonly",width=18)
        cb_pesa.pack(side="left",padx=4)

        def update_pesas(*a):
            pesas=BALANZAS_CFG[bal_var.get()]["pesas"]
            labels=[p[1] for p in pesas]
            cb_pesa["values"]=labels
            pesa_var.set(labels[0])
        bal_var.trace_add("write",lambda *a: update_pesas())
        update_pesas()

        resultado={"sel":None}

        def ok():
            bkey=bal_var.get()
            plbl=pesa_var.get()
            pesas=BALANZAS_CFG[bkey]["pesas"]
            pesa=next((p for p in pesas if p[1]==plbl),None)
            if pesa:
                resultado["sel"]=(bkey,pesa)
            win.destroy()

        def cancelar():
            win.destroy()

        btn_row=tk.Frame(win,bg=C["surface"])
        btn_row.pack(pady=(10,4))
        tk.Button(btn_row,text="  Aceptar  ",
            bg=C["accent"],fg=C["bg"],
            font=F["h3"],relief="flat",pady=8,
            command=ok).pack(side="left",padx=6)
        tk.Button(btn_row,text="Cancelar",
            bg=C["surface3"],fg=C["text2"],
            font=F["body"],relief="flat",pady=8,padx=10,
            command=cancelar).pack(side="left",padx=6)

        win.bind("<Return>",lambda e: ok())
        win.bind("<Escape>",lambda e: cancelar())
        self.root.wait_window(win)
        return resultado["sel"]

    def _exportar_xlsx(self):
        sel=self._get_seleccion()
        if sel:
            exportar_excel(sel[0],sel[1],self.hist,self.root)

    def _exportar_pdf(self):
        sel=self._get_seleccion()
        if sel:
            exportar_pdf(sel[0],sel[1],self.hist,self.root)

    def _tick(self):
        self.lbl_reloj.config(
            text=datetime.now().strftime("%Y-%m-%d  %H:%M:%S"))
        self.root.after(1000,self._tick)


if __name__=="__main__":
    root=tk.Tk()
    App(root)
    root.mainloop()
